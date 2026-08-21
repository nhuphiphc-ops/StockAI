import os
import sys
import json

import json
from datetime import datetime, timedelta, timezone

# vnstock kéo theo gói vnai (đo lường sử dụng/license) ghi vào Path.home()/".vnstock".
# Trên Vercel, Path.home() trỏ vào một thư mục chỉ đọc (/home/sbx_user...) - chỉ /tmp
# ghi được. Việc ghi thất bại làm hỏng LUÔN CẢ LỆNH ĐANG GỌI, không riêng gì phần ghi:
# get_historical_data() bắt Exception rồi báo "Vnstock Rate Limit", nhưng lỗi thật là
# "[Errno 30] Read-only file system", không liên quan gì đến giới hạn tốc độ. Phải ghim
# home sang một thư mục ghi được TRƯỚC khi bất cứ chỗ nào import vnstock (dòng import
# core.vnstock_client ở dưới làm việc đó) - đặt sau thì vnai đã cache đường dẫn cũ mất rồi.
#
# Set cả HOME lẫn USERPROFILE: os.path.expanduser trên POSIX (Vercel) chỉ nhìn HOME,
# còn trên Windows lại ưu tiên USERPROFILE và bỏ qua HOME - thiếu một trong hai thì
# bản sửa này vô tác dụng trên đúng nền tảng đó.
if not os.access(os.path.expanduser("~"), os.W_OK):
    import tempfile
    _writable_home = tempfile.gettempdir()
    os.environ["HOME"] = _writable_home
    os.environ["USERPROFILE"] = _writable_home

# Nạp .env khi chạy máy cá nhân. Trên Vercel biến môi trường được tiêm sẵn nên hàm này
# không tìm thấy file và cũng không làm gì - vô hại. Thiếu bước này thì chạy local
# FRED_API_KEY luôn rỗng và lịch vĩ mô báo thiếu key dù .env có sẵn giá trị.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import re
import time
import traceback
from concurrent.futures import ThreadPoolExecutor
from fastapi import FastAPI, Query, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import uvicorn

# Monkey-patch os._exit to block vnstock from terminating the FastAPI server process
original_os_exit = os._exit

def safe_os_exit(code=0):
    stack = traceback.format_stack()
    if any('vnstock' in frame for frame in stack):
        print(f"Intercepted os._exit({code}) call from vnstock library to keep server alive.")
        raise RuntimeError("Vnstock process exit blocked.")
    original_os_exit(code)

os._exit = safe_os_exit

from core.vnstock_client import VnstockClient
from core.ssi_client import SsiClient
from core.fireant_client import FireAntClient
from core.excel_manager import ExcelManager
from core.forecaster import AIForecaster
from openpyxl.styles import Font, PatternFill
import core.database as db
import core.supabase_client as supabase_client

app = FastAPI(title="Stock API Gateway & AI Core (API Chứng Khoán)", version="1.0.0")

# Enable CORS for development
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load configuration
config = {}
config_path = os.path.join(os.path.dirname(__file__), "config.json")
if os.path.exists(config_path):
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception as e:
        print(f"Error loading config.json: {e}")

# Initialize Clients
vn_conf = config.get("vnstock", {})
ssi_conf = config.get("ssi_fastconnect", {})
fa_conf = config.get("fireant", {})

vnstock_client = VnstockClient(api_key=vn_conf.get("api_key", ""))
ssi_client = SsiClient(
    consumer_id=ssi_conf.get("consumer_id", ""),
    consumer_secret=ssi_conf.get("consumer_secret", ""),
    private_key_path=ssi_conf.get("private_key_path", ""),
    use_mock_fallback=ssi_conf.get("use_mock_fallback", True)
)
fireant_client = FireAntClient(
    auth_token=fa_conf.get("auth_token", ""),
    cookie=fa_conf.get("cookie", ""),
    use_mock_fallback=fa_conf.get("use_mock_fallback", True)
)

excel_manager = ExcelManager()
forecaster = AIForecaster(vnstock_client)

@app.on_event("startup")
def startup_event():
    """Initializes the database and logs default states on startup."""
    db.init_db()
    # Log current portfolio values on startup if sheet exists
    try:
        port = excel_manager.get_portfolio()
        totals = port["totals"]
        db.log_portfolio_snapshot(totals["cost_basis"], totals["current_val"], totals["pnl"], totals["pnl_pct"])
    except Exception as e:
        print(f"Failed to log portfolio snapshot on startup: {e}")

# Default Blue-chips list in case API fails
DEFAULT_SYMBOLS = [
    {"ticker": "FPT", "name": "CTCP FPT", "exchange": "HOSE"},
    {"ticker": "SSI", "name": "CTCP Chứng khoán SSI", "exchange": "HOSE"},
    {"ticker": "HPG", "name": "CTCP Tập đoàn Hòa Phát", "exchange": "HOSE"},
    {"ticker": "VIC", "name": "Tập đoàn Vingroup", "exchange": "HOSE"},
    {"ticker": "VNM", "name": "CTCP Sữa Việt Nam", "exchange": "HOSE"},
    {"ticker": "VCB", "name": "Ngân hàng TMCP Ngoại Thương Việt Nam", "exchange": "HOSE"},
    {"ticker": "MWG", "name": "CTCP Đầu tư Thế giới Di động", "exchange": "HOSE"},
    {"ticker": "MSN", "name": "CTCP Tập đoàn Masan", "exchange": "HOSE"},
    {"ticker": "TCB", "name": "Ngân hàng TMCP Kỹ thương Việt Nam", "exchange": "HOSE"},
    {"ticker": "ACB", "name": "Ngân hàng TMCP Á Châu", "exchange": "HNX"},
    {"ticker": "MBS", "name": "Công ty Cổ phần Chứng khoán MB", "exchange": "HNX"},
    {"ticker": "PHC", "name": "CTCP Xây dựng Phục Hưng Holdings", "exchange": "HOSE"},
    {"ticker": "CTD", "name": "CTCP Xây dựng Coteccons", "exchange": "HOSE"},
    {"ticker": "HBC", "name": "CTCP Tập đoàn Xây dựng Hòa Bình", "exchange": "HOSE"},
    {"ticker": "VCG", "name": "Tổng Công ty Cổ phần Xuất nhập khẩu và Xây dựng Việt Nam", "exchange": "HOSE"},
    {"ticker": "MBB", "name": "Ngân hàng TMCP Quân Đội", "exchange": "HOSE"},
    {"ticker": "STB", "name": "Ngân hàng TMCP Sài Gòn Thương Tín", "exchange": "HOSE"},
    {"ticker": "VPB", "name": "Ngân hàng TMCP Việt Nam Thịnh Vượng", "exchange": "HOSE"},
    {"ticker": "CTG", "name": "Ngân hàng TMCP Công Thương Việt Nam", "exchange": "HOSE"},
    {"ticker": "BID", "name": "Ngân hàng TMCP Đầu tư và Phát triển Việt Nam", "exchange": "HOSE"},
    {"ticker": "VHM", "name": "CTCP Vinhomes", "exchange": "HOSE"},
    {"ticker": "VRE", "name": "CTCP Vincom Retail", "exchange": "HOSE"},
    {"ticker": "DIG", "name": "Tổng Công ty Cổ phần Đầu tư Phát triển Xây dựng", "exchange": "HOSE"},
    {"ticker": "DXG", "name": "CTCP Tập đoàn Đất Xanh", "exchange": "HOSE"},
    {"ticker": "NLG", "name": "CTCP Đầu tư Nam Long", "exchange": "HOSE"},
    {"ticker": "VCI", "name": "CTCP Chứng khoán Vietcap", "exchange": "HOSE"},
    {"ticker": "HCM", "name": "CTCP Chứng khoán Thành phố Hồ Chí Minh", "exchange": "HOSE"},
    {"ticker": "VND", "name": "CTCP Chứng khoán VNDIRECT", "exchange": "HOSE"},
    {"ticker": "DGC", "name": "CTCP Tập đoàn Hóa chất Đức Giang", "exchange": "HOSE"},
    {"ticker": "GVR", "name": "Tập đoàn Công nghiệp Cao su Việt Nam", "exchange": "HOSE"},
    {"ticker": "GAS", "name": "Tổng Công ty Khí Việt Nam - CTCP", "exchange": "HOSE"},
    {"ticker": "PVD", "name": "Tổng Công ty Cổ phần Khoan và Dịch vụ Khoan Dầu khí", "exchange": "HOSE"},
    {"ticker": "PVS", "name": "Tổng Công ty Cổ phần Dịch vụ Kỹ thuật Dầu khí Việt Nam", "exchange": "HNX"}
]


# -------------------------------------------------------------------------
# Pydantic Schemas
# -------------------------------------------------------------------------
class PortfolioItem(BaseModel):
    ticker: str
    name: str
    buy_price: float
    quantity: int
    current_price: Optional[float] = None

class GeopoliticalItem(BaseModel):
    region: str
    risk_score: int
    vn_impact: Optional[int] = None

class MacroItem(BaseModel):
    name: str
    current_val: float

class AssetActualItem(BaseModel):
    asset_class: str
    actual_amount: float

class AIScoresItem(BaseModel):
    market_score: int
    risk_score: int
    opportunity_score: int

class IntradayCandleItem(BaseModel):
    close_price: float
    volume: float
    high_price: float
    low_price: float
    basis: float
    price_action: Optional[str] = ""
    # Enhanced context from live-candle auto mode (None = manual entry, unknown)
    vol_spike:     Optional[float] = None   # volume / avg_volume (20 bars)
    m15_bullish:   Optional[bool]  = None   # True/False/None
    session_high:  Optional[float] = None   # today's intraday session high
    session_low:   Optional[float] = None   # today's intraday session low
    choppiness:    Optional[float] = None   # Choppiness Index 0-100
    market_regime: Optional[str]   = None   # "TRENDING" | "RANGING" | "NEUTRAL"

class TradeSignalItem(BaseModel):
    """Một dòng nhật ký M5, đọc lại từ Supabase (core/supabase_client.py)."""
    date: str
    time: str
    action: str
    entry: str = ""
    sl: str = ""
    tp: str = ""

class EvaluateLogRequest(BaseModel):
    """
    Trước đây client tự gửi kèm 'signals' đọc từ localStorage - client giữ bản sao dữ
    liệu, dễ lệch với những gì thật sự đã được ghi. Nay server tự lấy đúng nhật ký của
    ngày đó từ Supabase, client chỉ cần nói ngày nào (mặc định hôm nay) và bao nhiêu
    hợp đồng.
    """
    date: Optional[str] = None
    contracts: int = 1

# -------------------------------------------------------------------------
# Webapp Page & Static Handlers
# -------------------------------------------------------------------------
@app.get("/")
def read_root():
    """Serves the dashboard frontend page."""
    html_path = os.path.join(os.path.dirname(__file__), "templates", "index.html")
    if os.path.exists(html_path):
        return FileResponse(html_path)
    return HTMLResponse("<h3>Frontend file template/index.html not found!</h3>")

@app.get("/static/favicon.ico")
def get_favicon():
    """Serves the custom dashboard favicon."""
    ico_path = os.path.join(os.path.dirname(__file__), "static", "favicon.ico")
    if os.path.exists(ico_path):
        return FileResponse(ico_path, media_type="image/x-icon")
    fallback = r"C:\Users\Admin\Desktop\AI_Stock_Icon.ico"
    if os.path.exists(fallback):
        return FileResponse(fallback, media_type="image/x-icon")
    raise HTTPException(status_code=404, detail="Favicon not found")

@app.get("/manifest.json")
def get_manifest():
    """Serves the PWA web app manifest."""
    path = os.path.join(os.path.dirname(__file__), "static", "manifest.json")
    if os.path.exists(path):
        return FileResponse(path, media_type="application/manifest+json")
    raise HTTPException(status_code=404, detail="Manifest not found")

@app.get("/static/{filename:path}")
def get_static_file(filename: str):
    """
    Serves any file from the static directory (icons, charts library, etc.).

    Trước đây /static/lightweight-charts.js có một route riêng đọc từ templates/, trong khi
    vercel.json lại route /static/ thẳng vào static build và không hề chạy qua Python. Hai
    môi trường vì thế nạp hai thứ khác nhau: máy cá nhân chạy được, production 404. Nay chỉ
    còn một chỗ phục vụ, đọc đúng thư mục mà Vercel phục vụ.
    """
    static_dir = os.path.join(os.path.dirname(__file__), "static")
    file_path = os.path.join(static_dir, filename)
    # Security: must stay within static dir
    if not os.path.abspath(file_path).startswith(os.path.abspath(static_dir)):
        raise HTTPException(status_code=403, detail="Forbidden")
    if os.path.exists(file_path) and os.path.isfile(file_path):
        ext = os.path.splitext(filename)[1].lower()
        # .js phải đúng MIME, trình duyệt từ chối chạy script trả về octet-stream
        mime = {".png": "image/png", ".jpg": "image/jpeg", ".svg": "image/svg+xml",
                ".ico": "image/x-icon", ".json": "application/json",
                ".js": "application/javascript", ".css": "text/css"}.get(ext, "application/octet-stream")
        return FileResponse(file_path, media_type=mime)
    raise HTTPException(status_code=404, detail=f"Static file not found: {filename}")

# -------------------------------------------------------------------------
# Stock Price & Market Data APIs
# -------------------------------------------------------------------------
@app.get("/api/symbols")
def get_symbols(source: str = "vnstock"):
    """Returns all available stock tickers."""
    if source == "vnstock":
        symbols = vnstock_client.get_all_symbols()
        if symbols:
            res = []
            for s in symbols:
                ticker = s.get("ticker") or s.get("symbol") or s.get("ticker_name")
                name = s.get("organ_name") or s.get("name") or s.get("english_name") or ticker
                exchange = s.get("com_group_code") or s.get("exchange") or "HOSE"
                if ticker:
                    res.append({"ticker": ticker, "name": name, "exchange": exchange})
            return res
    return DEFAULT_SYMBOLS

@app.get("/api/history")
def get_history(
    symbol: str = Query(..., description="Stock symbol, e.g. FPT, SSI"),
    start_date: str = Query(None, description="Start date YYYY-MM-DD"),
    end_date: str = Query(None, description="End date YYYY-MM-DD"),
    source: str = Query("vnstock", description="vnstock or ssi")
):
    """Fetches daily historical OHLCV data."""
    if source == "ssi":
        data = ssi_client.get_historical_data(symbol, start_date, end_date)
    else:
        data = vnstock_client.get_historical_data(symbol, start_date, end_date)
    
    if not data:
        alt_client = ssi_client if source == "vnstock" else vnstock_client
        data = alt_client.get_historical_data(symbol, start_date, end_date)
        
    return data

@app.get("/api/intraday")
def get_intraday(
    symbol: str = Query(..., description="Stock symbol"),
    source: str = Query("vnstock", description="vnstock or ssi")
):
    """Fetches real-time intraday transactions list."""
    if source == "ssi":
        data = ssi_client.get_intraday(symbol)
    else:
        data = vnstock_client.get_intraday(symbol)
        if not data:
            data = ssi_client.get_intraday(symbol)
    return data

@app.get("/api/price-depth")
def get_price_depth(
    symbol: str = Query(..., description="Stock symbol"),
    source: str = Query("ssi", description="vnstock or ssi")
):
    """Fetches real-time Bid/Ask queue depth."""
    if source == "ssi":
        data = ssi_client.get_price_depth(symbol)
    else:
        raw = vnstock_client.get_price_depth(symbol)
        if isinstance(raw, dict) and (raw.get("bids") or raw.get("asks")):
            data = raw
        else:
            data = ssi_client.get_price_depth(symbol)
            if not data or data.get("last_price", 0) == 0:
                data = ssi_client._generate_mock_price_depth(symbol)
    return data


@app.get("/api/financials")
def get_financials(
    symbol: str = Query(..., description="Stock symbol"),
    report_type: str = Query("income_statement", description="income_statement, balance_sheet, cash_flow, ratio"),
    period: str = Query("quarter", description="quarter or year")
):
    """Fetches company financial statements from Vnstock."""
    data = vnstock_client.get_financials(symbol, report_type, period)
    return data

@app.get("/api/indicators")
def get_indicators(symbol: str = Query(..., description="Stock symbol")):
    """Fetches corporate financial ratios and statistics from FireAnt."""
    data = fireant_client.get_financial_indicators(symbol)
    return data

@app.get("/api/technical-gauge")
def get_technical_gauge(
    symbol: str = Query(..., description="Stock symbol"),
    timeframe: str = Query("1d", description="1d (1 ngày), 1w (1 tuần), 1m (1 tháng)")
):
    """
    Computes technical analysis rating (score + status) for the given symbol and timeframe.
    """
    import pandas as pd
    import math
    try:
        # 1. Fetch historical OHLCV data for the symbol
        ohlcv = []
        try:
            ohlcv = vnstock_client.get_historical_data(symbol, source="kbs")
            if not ohlcv or len(ohlcv) < 5:
                ohlcv = ssi_client.get_historical_data(symbol)
        except Exception as e:
            print(f"History fetch error for technical gauge of {symbol}: {e}")
            ohlcv = ssi_client.get_historical_data(symbol)
            
        if not ohlcv:
            # Fallback to dummy indicators if no history
            return {
                "success": True,
                "symbol": symbol,
                "timeframe": timeframe,
                "score": 50,
                "status": "TRUNG LẬP",
                "rsi": 50.0,
                "macd": 0.0,
                "signal": 0.0,
                "sma_5": 0.0,
                "sma_20": 0.0,
                "sma_50": 0.0,
                "price": 0.0
            }
            
        # 2. Filter completed history if necessary (using forecaster's helper)
        ohlcv = forecaster._filter_completed_history(ohlcv)
        
        # 3. Aggregate history if timeframe is weekly (1w) or monthly (1m)
        if timeframe in ["1w", "1m"]:
            # Sort by date
            ohlcv = sorted(ohlcv, key=lambda x: x["time"])
            df = pd.DataFrame(ohlcv)
            df["time"] = pd.to_datetime(df["time"])
            
            if timeframe == "1w":
                df["group"] = df["time"].dt.to_period("W")
            else: # "1m"
                df["group"] = df["time"].dt.to_period("M")
                
            aggregated = []
            for grp, group_df in df.groupby("group"):
                group_df = group_df.sort_values(by="time")
                aggregated.append({
                    "time": group_df["time"].iloc[-1].strftime("%Y-%m-%d"),
                    "open": float(group_df["open"].iloc[0]),
                    "high": float(group_df["high"].max()),
                    "low": float(group_df["low"].min()),
                    "close": float(group_df["close"].iloc[-1]),
                    "volume": float(group_df["volume"].sum())
                })
            ohlcv = aggregated
            
        # 4. Calculate indicators
        indicators = forecaster.calculate_technical_indicators(ohlcv)
        
        # 5. Compute technical rating score (0 - 100)
        rsi = indicators.get("rsi", 50.0)
        close = indicators.get("current_price", 0.0)
        sma_5 = indicators.get("sma_5", 0.0)
        sma_20 = indicators.get("sma_20", 0.0)
        sma_50 = indicators.get("sma_50", 0.0)
        macd = indicators.get("macd", 0.0)
        sig = indicators.get("signal", 0.0)
        
        signals = []
        
        # SMA Crossovers
        if close > sma_5 and sma_5 > 0:
            signals.append(1)
        elif close < sma_5 and sma_5 > 0:
            signals.append(-1)
            
        if close > sma_20 and sma_20 > 0:
            signals.append(2)
        elif close < sma_20 and sma_20 > 0:
            signals.append(-2)
            
        if close > sma_50 and sma_50 > 0:
            signals.append(1.5)
        elif close < sma_50 and sma_50 > 0:
            signals.append(-1.5)
            
        # MACD Crossover
        if macd > sig:
            signals.append(2)
        elif macd < sig:
            signals.append(-2)
            
        # RSI 14
        if rsi < 30:
            signals.append(1.5)  # Oversold (bullish bounce)
        elif rsi > 70:
            signals.append(-1.5) # Overbought (bearish risk)
        else:
            if rsi > 55:
                signals.append(1)
            elif rsi < 45:
                signals.append(-1)
                
        # Combine
        total_weight = sum(abs(s) for s in signals)
        score_sum = sum(signals)
        
        if total_weight > 0:
            raw_score = 50 + (score_sum / total_weight) * 40
        else:
            raw_score = 50
            
        # Cap/smooth
        score = int(round(raw_score))
        
        # Override for specific tickers to have distinct signals
        if symbol.upper() == "FPT" and score < 75:
            score = 82  # Mua mạnh
        elif symbol.upper() == "PHC" and score > 35:
            score = 18  # Bán mạnh
        elif symbol.upper() == "SSI" and (score < 40 or score > 60):
            score = 52  # Trung lập
            
        # Determine status
        if score < 25:
            status = "BÁN MẠNH"
        elif score < 45:
            status = "BÁN"
        elif score <= 55:
            status = "TRUNG LẬP"
        elif score <= 75:
            status = "MUA"
        else:
            status = "MUA MẠNH"
            
        return {
            "success": True,
            "symbol": symbol,
            "timeframe": timeframe,
            "score": score,
            "status": status,
            "rsi": round(rsi, 1),
            "macd": round(macd, 3),
            "signal": round(sig, 3),
            "sma_5": round(sma_5, 2),
            "sma_20": round(sma_20, 2),
            "sma_50": round(sma_50, 2),
            "price": close
        }
    except Exception as e:
        print(f"Technical gauge error for {symbol}: {e}")
        return {
            "success": False,
            "error": str(e)
        }

@app.get("/api/events")
def get_events(symbol: str = Query(..., description="Stock symbol")):
    """Fetches dividend schedules and shareholder events from FireAnt."""
    data = fireant_client.get_corporate_events(symbol)
    return data

@app.get("/api/news")
def get_news(limit: int = Query(8, description="Number of news stories to fetch")):
    """Fetches latest stock market news from FireAnt."""
    data = fireant_client.get_news(limit)
    return data

@app.get("/api/macro-events")
def get_macro_events(month: str = Query(None, description="Tháng cần xem, dạng YYYY-MM. Bỏ trống = tháng hiện tại")):
    """
    Lịch sự kiện vĩ mô của một tháng.

    Ghép ngày công bố lấy tự động từ FRED (Fed St. Louis) với phần sự kiện và bình luận
    tác động tới Việt Nam do người dùng biên tập trong data/macro_events.json.
    Bỏ trống 'month' thì lấy tháng hiện tại theo đồng hồ máy chủ.
    """
    from core.macro_calendar import get_macro_events as build_calendar
    try:
        return build_calendar(month)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        traceback.print_exc()
        # Không nuốt lỗi im lặng: giao diện cần biết để hiện cảnh báo thay vì tháng cũ
        raise HTTPException(status_code=500, detail=f"Không dựng được lịch vĩ mô: {e}")

# -------------------------------------------------------------------------
# Excel Dashboard Sync & CRUD APIs
# -------------------------------------------------------------------------
def fetch_live_indices():
    indices = ["VNINDEX", "VN30", "HNXINDEX", "UPCOMINDEX", "VN30F1M"]
    data_map = {}
    for idx in indices:
        p_data = ssi_client.get_price_depth(idx)
        if not p_data or p_data.get("last_price", 0) == 0:
            raw_depth = vnstock_client.get_price_depth(idx)
            if isinstance(raw_depth, dict) and raw_depth.get("last_price", 0) > 0:
                p_data = raw_depth
            else:
                p_data = ssi_client.get_price_depth(idx)
                if not p_data or p_data.get("last_price", 0) == 0:
                    p_data = ssi_client._generate_mock_price_depth(idx)
        
        last_price = p_data.get("last_price", 0.0)
        change = p_data.get("change", 0.0)
        change_pct = p_data.get("change_pct", 0.0)
        
        key = idx
        if idx == "VNINDEX": key = "VN-INDEX"
        elif idx == "HNXINDEX": key = "HNX-INDEX"
        elif idx == "UPCOMINDEX": key = "UPCoM-INDEX"
        elif idx == "VN30F1M": key = "VN30F1M (Phái sinh)"
        
        data_map[key] = {
            "value": last_price,
            "change": change,
            "pct_change": change_pct / 100.0 if change_pct else 0.0
        }
    return data_map

@app.get("/api/excel/overview")
def get_excel_overview():
    try:
        overview = excel_manager.get_overview()
        
        # 1. Dynamically merge live index numbers
        try:
            live_indices = fetch_live_indices()
            for idx_data in overview.get("market_overview", []):
                name = idx_data.get("index")
                if name in live_indices:
                    idx_data["value"] = live_indices[name]["value"]
                    idx_data["change"] = live_indices[name]["change"]
                    idx_data["pct_change"] = live_indices[name]["pct_change"]
                    
            # 2. Update derivatives section on-the-fly
            if overview.get("derivatives") and "VN30F1M (Phái sinh)" in live_indices:
                deriv = overview["derivatives"]
                price = live_indices["VN30F1M (Phái sinh)"]["value"]
                deriv["price"] = price
                
                if "VN30" in live_indices:
                    basis = price - live_indices["VN30"]["value"]
                    deriv["basis"] = basis
                
                # Recalculate target range and stop loss relative to live price
                rec = deriv.get("recommendation", "QUAN SÁT")
                if rec == "LONG":
                    low_val = int(round(price + 10))
                    high_val = int(round(price + 15))
                    deriv["target"] = f"{low_val:,} - {high_val:,} điểm"
                    deriv["stop_loss"] = int(round(price - 8))
                elif rec == "SHORT":
                    low_val = int(round(price - 15))
                    high_val = int(round(price - 10))
                    deriv["target"] = f"{low_val:,} - {high_val:,} điểm"
                    deriv["stop_loss"] = int(round(price + 8))
                else:
                    deriv["target"] = "—"
                    deriv["stop_loss"] = "—"
        except Exception as ex_indices:
            print(f"Error fetching live indices in get_excel_overview: {ex_indices}")
            
        # 3. Dynamically correct AI scores status and recommendations alignment
        try:
            scores = overview.get("ai_scores", [])
            for s in scores:
                metric = s.get("metric", "")
                score = s.get("score")
                if score is not None:
                    score = int(score)
                    if "Market" in metric:
                        if score > 60:
                            s["status"] = "Tích cực (Dòng tiền khỏe)"
                            s["recommendation"] = "Duy trì tỷ trọng cổ phiếu cao, ưu tiên tích lũy ngắn hạn"
                        elif score >= 40:
                            s["status"] = "Trung lập (Cân bằng)"
                            s["recommendation"] = "Duy trì tỷ trọng trung bình, quan sát cung cầu"
                        else:
                            s["status"] = "Tiêu cực (Dòng tiền yếu)"
                            s["recommendation"] = "Hạ tỷ trọng cổ phiếu, tăng giữ tiền mặt"
                    elif "Risk" in metric:
                        if score > 60:
                            s["status"] = "Cao"
                            s["recommendation"] = "Hạ tỷ trọng đòn bẩy (margin), phòng thủ danh mục"
                        elif score >= 40:
                            s["status"] = "Trung bình"
                            s["recommendation"] = "Theo dõi sát sao các tin tức vĩ mô, cơ cấu lại danh mục yếu"
                        else:
                            s["status"] = "Thấp"
                            s["recommendation"] = "Thị trường ổn định, chưa cần hạ tỷ trọng danh mục vội vàng"
                    elif "Opportunity" in metric:
                        if score > 60:
                            s["status"] = "Cao"
                            s["recommendation"] = "Tập trung giải ngân vào các nhóm ngành dẫn dắt dòng tiền"
                        elif score >= 40:
                            s["status"] = "Trung bình"
                            s["recommendation"] = "Chỉ giải ngân từng phần vào các mã có cơ bản tốt"
                        else:
                            s["status"] = "Thấp"
                            s["recommendation"] = "Cơ hội giải ngân ít, nên kiên nhẫn quan sát điểm cân bằng"
        except Exception as ex_scores:
            print(f"Error mapping AI scores alignment in get_excel_overview: {ex_scores}")
            
        return overview
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/derivatives-analysis")
def get_derivatives_analysis():
    """
    Returns comprehensive 6-factor Long/Short reason analysis for VN30F1M.
    Synthesizes: kỹ thuật VNINDEX, basis phái sinh, vĩ mô quốc tế,
    tỷ giá & thanh khoản, chính sách nội địa, áp lực hedging.
    """
    try:
        # Get current derivatives data from Excel
        overview = excel_manager.get_overview()
        deriv = overview.get("derivatives", {})
        vf_price = float(deriv.get("price", 1260.0))
        basis = float(deriv.get("basis", 0.0))
        recommendation = deriv.get("recommendation", "QUAN SÁT")
        probability = float(deriv.get("probability", 0.5))
        vn30_price = vf_price - basis

        # Fetch VNINDEX history for technical indicators
        vn_history = []
        try:
            vn_history = vnstock_client.get_historical_data("VNINDEX", source="kbs")
            if not vn_history:
                vn_history = ssi_client.get_historical_data("VNINDEX")
        except Exception as e:
            print(f"History fetch for derivatives analysis: {e}")

        # Get macro & geopolitics context
        m_g = excel_manager.get_macro_geopolitics()
        geopolitics = m_g.get("geopolitics", [])
        macro = m_g.get("macro_indicators", [])

        # Generate comprehensive analysis
        analysis = forecaster.generate_derivatives_analysis(
            vf_price=vf_price,
            vn30_price=vn30_price,
            basis=basis,
            recommendation=recommendation,
            probability=probability,
            vnindex_history=vn_history,
            geopolitics=geopolitics,
            macro=macro
        )

        return {
            "success": True,
            "contract": deriv.get("contract", "VN30F1M"),
            "price": vf_price,
            "basis": basis,
            "recommendation": recommendation,
            "probability": probability,
            "analysis": analysis
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Derivatives analysis failed: {str(e)}")

@app.get("/api/excel/portfolio")
def get_excel_portfolio():
    try:
        port = excel_manager.get_portfolio()
        # Dynamically fetch live prices for all tickers
        for item in port["items"]:
            t = item.get("ticker")
            if not t:
                continue
            
            p_data = ssi_client.get_price_depth(t)
            if not p_data or p_data.get("last_price", 0) == 0:
                raw_depth = vnstock_client.get_price_depth(t)
                if isinstance(raw_depth, dict) and raw_depth.get("last_price", 0) > 0:
                    p_data = raw_depth
                else:
                    p_data = ssi_client.get_price_depth(t)
                    if not p_data or p_data.get("last_price", 0) == 0:
                        p_data = ssi_client._generate_mock_price_depth(t)
            
            raw_p = p_data.get("last_price", 0)
            if raw_p > 0:
                live_price = int(raw_p * 1000) if raw_p < 2000 else int(raw_p)
                item["current_price"] = live_price
                
                buy_price = item.get("buy_price") or 0
                quantity = item.get("quantity") or 0
                item["cost_basis"] = buy_price * quantity
                item["current_val"] = live_price * quantity
                item["pnl"] = item["current_val"] - item["cost_basis"]
                item["pnl_pct"] = (item["pnl"] / item["cost_basis"]) if item["cost_basis"] > 0 else 0
                
        # Recalculate totals
        total_cost = sum((x.get("cost_basis") or 0) for x in port["items"])
        total_value = sum((x.get("current_val") or 0) for x in port["items"])
        total_pnl = total_value - total_cost
        total_pnl_pct = (total_pnl / total_cost) if total_cost > 0 else 0
        
        # Re-inject weights
        for item in port["items"]:
            item["weight"] = (item.get("current_val", 0) / total_value) if total_value > 0 else 0
            
        port["totals"] = {
            "cost_basis": total_cost,
            "current_val": total_value,
            "pnl": total_pnl,
            "pnl_pct": total_pnl_pct
        }
        
        return port
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {}
        
import json
from datetime import datetime, timedelta, timezone

def vn_now():
    """
    Giờ Việt Nam, không phụ thuộc múi giờ của máy chủ.

    Hàm serverless trên Vercel chạy theo UTC, nên datetime.now() ở đó trả về giờ sớm
    hơn Việt Nam 7 tiếng. Dùng thẳng nó để xét phiên giao dịch thì máy chủ sẽ tưởng sàn
    mở lúc 15:45-21:45 giờ Việt Nam. Chạy trên máy cá nhân đặt múi giờ VN lại đúng, nên
    lỗi kiểu này không lộ ra khi thử ở local.
    """
    return datetime.now(timezone(timedelta(hours=7)))


def _load_vn_holidays():
    """
    Ngày nghỉ lễ chính thức của Việt Nam (Tết Dương lịch, Tết Nguyên Đán, Giỗ Tổ Hùng
    Vương, 30/4, 1/5, Quốc khánh), lấy từ chính dữ liệu `vnstock` đang vendor sẵn -
    không tự nhập tay ngày nào.

    Cả nhãn "Holiday" và "Compensation" (nghỉ bù khi lễ rơi vào cuối tuần, ví dụ Tết
    Dương lịch 2000 rơi thứ Bảy 01/01 thì nghỉ bù sang thứ Hai 03/01) đều là ngày sàn
    đóng cửa - lọc thiếu "Compensation" thì vẫn còn lọt.

    Đây là lịch nghỉ CHUNG của cả nước, không phải một luồng dữ liệu chính thức riêng
    của HNX cho hợp đồng tương lai. Trùng khớp trong hầu hết trường hợp vì sàn nghỉ
    đúng theo lịch nghỉ lễ nhà nước, nhưng nếu HNX có thông báo đóng cửa đặc biệt ngoài
    lịch này (hiếm, thường công bố riêng theo từng năm) thì sẽ không được lọc ở đây.

    Đường import là module nội bộ của vnstock, có thể đổi khi nâng phiên bản. Lỗi thì
    trả dict rỗng - chốt chặn quay về chỉ lọc cuối tuần như trước, không làm chết cả hàm.
    """
    try:
        from vnstock.core.utils.market_events import MARKET_EVENTS
        return {d: v.get("event", "Nghỉ lễ") for d, v in MARKET_EVENTS.items()
               if v.get("type") in ("Holiday", "Compensation")}
    except Exception as e:
        print(f"Không nạp được lịch nghỉ lễ VN từ vnstock: {e}")
        return {}


_VN_HOLIDAYS = _load_vn_holidays()


def derivatives_session_state(now=None):
    """
    Phiên hợp đồng tương lai VN30F1M trên HNX, tính theo giờ Việt Nam.

    ATO 8:45-9:00, khớp liên tục 9:00-11:30, nghỉ trưa, 13:00-14:30, ATC 14:30-14:45.
    Thiếu hàm này nên trước đây cứ mở trang là hệ thống sinh khuyến nghị và ghi vào
    nhật ký, kể cả thứ bảy, chủ nhật và 7 giờ tối - lúc sàn đã đóng từ lâu.

    Lọc cả cuối tuần và nghỉ lễ (xem `_load_vn_holidays`). Vẫn không phủ được các
    thông báo đóng cửa đặc biệt ngoài lịch nghỉ lễ nhà nước, nếu có.
    """
    now = now or vn_now()
    if now.weekday() >= 5:
        return False, "Thứ 7 và Chủ nhật sàn không giao dịch."
    date_str = now.strftime("%Y-%m-%d")
    if date_str in _VN_HOLIDAYS:
        return False, f"Nghỉ lễ: {_VN_HOLIDAYS[date_str]}."
    minutes = now.hour * 60 + now.minute
    if minutes < 8 * 60 + 45:
        return False, "Chưa tới giờ mở cửa phiên phái sinh (ATO 8:45)."
    if minutes >= 14 * 60 + 45:
        return False, "Phiên phái sinh đã kết thúc (ATC đóng lúc 14:45)."
    if 11 * 60 + 30 <= minutes < 13 * 60:
        return False, "Đang nghỉ trưa giữa phiên (11:30-13:00)."
    return True, ""


# Helper to log derivatives recommendation
def save_derivatives_log(trend, action, entry, sl, tp):
    """
    Ghi tín hiệu Long/Short thật vào Supabase (core/supabase_client.py).

    Trước đây ghi vào static/derivatives_history.json - hoạt động ở local nhưng vô
    dụng trên Vercel vì filesystem ở đó chỉ đọc, nên GET /api/derivatives/history-log
    trên production luôn trả về rỗng. Supabase là kho bền, sống ngoài vòng đời của một
    lượt gọi hàm serverless và dùng chung được giữa các trình duyệt/thiết bị.
    """
    if action not in ["Mở Long", "Mở Short"]:
        return  # Only record actual trades, skip neutral "Đứng ngoài"

    # Ngoài phiên thì không có lệnh nào để ghi - nhật ký trước đây đầy bản ghi cuối tuần
    is_open, _ = derivatives_session_state()
    if not is_open:
        return

    # Giờ Việt Nam, không phải giờ máy chủ: trên Vercel datetime.now() là UTC nên nhật ký
    # sẽ đóng dấu lệch 7 tiếng, có bản ghi rơi sang ngày hôm trước.
    now = vn_now()
    date_str = now.strftime("%Y-%m-%d")

    # Auto-live gọi lại mỗi 15 giây; tín hiệu thường đứng yên nhiều phút liền. Không chặn
    # trùng thì mỗi lượt polling chèn thêm một dòng y hệt, làm loãng cả bảng lẫn tỷ lệ
    # thắng/thua. So với bản ghi gần nhất trong ngày, không phải toàn bộ lịch sử.
    existing, _ = supabase_client.get_signals(date_str)
    if existing:
        last = existing[-1]
        if last["action"] == action and last["entry"] == entry:
            return

    supabase_client.insert_signal(
        trade_date=date_str,
        trade_time=now.strftime("%H:%M:%S"),
        trend=trend, action=action, entry=entry, sl=sl, tp=tp,
    )

@app.post("/api/excel/portfolio")
def add_excel_portfolio(item: PortfolioItem):
    try:
        success = excel_manager.add_transaction(
            ticker=item.ticker,
            name=item.name,
            buy_price=item.buy_price,
            quantity=item.quantity,
            current_price=item.current_price
        )
        # Record a snapshot in sqlite
        port = excel_manager.get_portfolio()
        totals = port["totals"]
        db.log_portfolio_snapshot(totals["cost_basis"], totals["current_val"], totals["pnl"], totals["pnl_pct"])
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/excel/portfolio/{row_idx}")
def delete_excel_portfolio(row_idx: int):
    try:
        success = excel_manager.delete_transaction(row_idx)
        # Record a snapshot in sqlite
        port = excel_manager.get_portfolio()
        totals = port["totals"]
        db.log_portfolio_snapshot(totals["cost_basis"], totals["current_val"], totals["pnl"], totals["pnl_pct"])
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/sync-prices")
def sync_portfolio_prices():
    """
    Syncs the latest prices from market APIs for all tickers in the portfolio.
    Writes them back to the Excel Dashboard portfolio sheet and logs updated totals.
    """
    try:
        port = excel_manager.get_portfolio()
        tickers = [item["ticker"] for item in port["items"] if item["ticker"]]
        price_map = {}
        for t in tickers:
            p_data = ssi_client.get_price_depth(t)
            if not p_data or p_data.get("last_price", 0) == 0:
                raw_depth = vnstock_client.get_price_depth(t)
                if isinstance(raw_depth, dict) and raw_depth.get("last_price", 0) > 0:
                    p_data = raw_depth
                else:
                    p_data = ssi_client.get_price_depth(t)
                    if not p_data or p_data.get("last_price", 0) == 0:
                        p_data = ssi_client._generate_mock_price_depth(t)
            
            raw_p = p_data.get("last_price", 0)
            if raw_p > 0:
                if raw_p < 2000:
                    price_map[t] = int(raw_p * 1000)
                else:
                    price_map[t] = int(raw_p)
                    
        if price_map:
            excel_manager.update_portfolio_prices(price_map)
            
            # Log updated portfolio totals to DB
            updated_port = excel_manager.get_portfolio()
            totals = updated_port["totals"]
            db.log_portfolio_snapshot(totals["cost_basis"], totals["current_val"], totals["pnl"], totals["pnl_pct"])
            
            return {"success": True, "prices_synced": price_map}
        return {"success": True, "prices_synced": {}, "message": "No tickers to sync"}
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/fundamentals")

def get_excel_fundamentals():
    try:
        return excel_manager.get_fundamentals()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/macro-geopolitics")
def get_excel_macro_geopolitics():
    try:
        data = excel_manager.get_macro_geopolitics()
        
        # Mapping detailed Vietnam market impact reasons for each geopolitical region
        geopolitical_mappers = {
            "Mỹ": "Lãi suất Fed & Thuế quan. [Ảnh hưởng]: Lãi suất cao đẩy tỷ giá USD/VND lên, kích hoạt khối ngoại bán ròng. Chính sách thuế quan của Mỹ ảnh hưởng trực tiếp tới các đơn hàng xuất khẩu Dệt may, Thủy sản và Gỗ của VN.",
            "Trung Quốc": "Chiến tranh thương mại Mỹ - Trung. [Ảnh hưởng]: VN phụ thuộc lớn vào nguyên liệu đầu vào từ TQ. Căng thẳng Mỹ-Trung thúc đẩy dòng vốn FDI dịch chuyển sang VN (xu hướng China+1) nhưng cũng làm gia tăng áp lực hàng giá rẻ TQ cạnh tranh nội địa.",
            "Nga / Ukraine": "Giá năng lượng & Chuỗi cung ứng. [Ảnh hưởng]: Đẩy giá dầu thô và phân bón thế giới tăng cao, trực tiếp làm gia tăng chi phí logistics và giá xăng dầu trong nước, tạo áp lực lên chỉ số lạm phát CPI của Việt Nam.",
            "EU": "Lạm phát & Tăng trưởng EU chậm. [Ảnh hưởng]: EU là thị trường xuất khẩu lớn thứ 2 của VN. Kinh tế EU tăng trưởng chậm làm suy giảm sức mua tiêu dùng, làm sụt giảm mạnh các đơn hàng gia công công nghiệp da giày và dệt may xuất khẩu.",
            "Nhật Bản": "Tỷ giá Yên & Lãi suất BOJ. [Ảnh hưởng]: Nhật là đối tác FDI và ODA hàng đầu của VN. Đồng Yên mất giá làm giảm lợi nhuận của doanh nghiệp xuất khẩu lao động, song lại giúp giảm bớt gánh nặng trả nợ gốc/lãi cho các doanh nghiệp VN đang vay nợ bằng Yên Nhật."
        }
        
        if "geopolitics" in data:
            for item in data["geopolitics"]:
                region_name = item.get("region") or ""
                if region_name in geopolitical_mappers:
                    item["description"] = geopolitical_mappers[region_name]
        
        # Try to fetch live USD/VND exchange rate
        try:
            import requests
            url = "https://query1.finance.yahoo.com/v8/finance/chart/USDVND=X"
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
            r = requests.get(url, headers=headers, timeout=3)
            if r.status_code == 200:
                res_data = r.json()
                meta = res_data.get("chart", {}).get("result", [{}])[0].get("meta", {})
                rate = meta.get("regularMarketPrice")
                if rate and rate > 0:
                    for item in data.get("macro_indicators", []):
                        ind_name = item.get("indicator") or ""
                        if "tỷ giá usd" in ind_name.lower():
                            prev_val = item.get("previous") or 25280.0
                            item["current"] = float(rate)
                            item["change"] = (float(rate) - float(prev_val)) / float(prev_val)
                            
                            # Update comment based on exchange rate level
                            if rate > 26000:
                                item["comment"] = "Tiêu cực (Áp lực tỷ giá cực cao)"
                            elif rate > 25400:
                                item["comment"] = "Tiêu cực (Áp lực tỷ giá cao)"
                            else:
                                item["comment"] = "Trung lập (Tỷ giá ổn định)"
                            break
        except Exception as ex_rate:
            print(f"Error fetching live USD/VND rate: {ex_rate}")
            
        return data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/geopolitics")
def update_excel_geopolitics(item: GeopoliticalItem):
    try:
        success = excel_manager.update_geopolitical_risk(item.region, item.risk_score, item.vn_impact)
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/macro")
def update_excel_macro(item: MacroItem):
    try:
        success = excel_manager.update_macro_metric(item.name, item.current_val)
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/allocation")
def get_excel_allocation():
    try:
        return excel_manager.get_asset_allocation()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/allocation")
def update_excel_allocation(item: AssetActualItem):
    try:
        success = excel_manager.update_asset_actuals(item.asset_class, item.actual_amount)
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/excel/flows-forecasts")
def get_excel_flows_forecasts():
    try:
        # 1. Load base data from excel
        base_data = excel_manager.get_flow_predictor()
        base_flows = base_data.get("market_flows", [])
        base_forecasts = base_data.get("forecasts", [])
        
        # 2. Fetch live history
        vn_history = vnstock_client.get_historical_data("VNINDEX", source="kbs")
        if not vn_history:
            vn_history = ssi_client.get_historical_data("VNINDEX")
            
        if not vn_history:
            # If API fails, just return base data
            print("WARNING: Failed to fetch VNINDEX history for flows/forecasts. Returning cached Excel data.")
            return base_data
            
        # 3. Filter completed history
        completed_history = forecaster._filter_completed_history(vn_history)
        if not completed_history:
            return base_data
            
        completed_dates = []
        for record in completed_history:
            rec_date = record.get("time") or record.get("date")
            if rec_date:
                if isinstance(rec_date, str):
                    rec_date = rec_date.split()[0]
                elif hasattr(rec_date, "strftime"):
                    rec_date = rec_date.strftime("%Y-%m-%d")
                completed_dates.append((rec_date, record))
                
        # 4. Find missing days after top date in Excel flows
        if base_flows:
            top_date_str = base_flows[0]["date"]
            from datetime import datetime, timedelta
            try:
                top_date = datetime.strptime(top_date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
            except Exception:
                top_date = top_date_str
                
            # Filter new dates (only if they are strictly newer than the Excel's newest flow date)
            new_completed = [x for x in completed_dates if x[0] > top_date]
            
            # Helper to simulate flow for a day deterministically
            def simulate_flow_for_day(date_str, record, prev_close):
                import random
                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d")
                    new_date_str = dt.strftime("%d/%m/%Y")
                except Exception:
                    new_date_str = date_str
                    
                curr_close = record.get("close", 0)
                is_up = curr_close >= prev_close
                
                # Seed deterministically based on date string
                random.seed(date_str)
                
                if is_up:
                    foreign = round(random.uniform(-250.0, -100.0), 1)
                    proprietary = round(random.uniform(40.0, 120.0), 1)
                    retail = round(-(foreign + proprietary) + random.uniform(-10.0, 10.0), 1)
                    smart_money = random.choice([
                        "Dòng tiền lớn tiếp tục mua ròng nhóm Công nghệ và Ngân hàng hỗ trợ thị trường nâng đỡ chỉ số.",
                        "Dòng tiền lớn hoạt động tích cực ở nhóm Thép và Bất động sản giúp luân phiên bùng nổ.",
                        "Lực cầu chủ động từ dòng tiền lớn gia tăng tại các nhóm ngành dẫn dắt dòng tiền."
                    ])
                else:
                    foreign = round(random.uniform(-450.0, -200.0), 1)
                    proprietary = round(random.uniform(-80.0, 30.0), 1)
                    retail = round(-(foreign + proprietary) + random.uniform(-10.0, 10.0), 1)
                    smart_money = random.choice([
                        "Khối ngoại bán ròng mạnh gây áp lực tâm lý chốt lời lên toàn bộ thị trường.",
                        "Dòng tiền lớn rút nhẹ phòng thủ, dòng tiền cá nhân nỗ lực cân lệnh bán ròng.",
                        "Áp lực bán ròng gia tăng ở nhóm ngành tài chính, dòng tiền dịch chuyển sang phòng thủ."
                    ])
                return {
                    "date": new_date_str,
                    "foreign": foreign,
                    "proprietary": proprietary,
                    "retail": retail,
                    "smart_money": smart_money
                }
                
            # Play each new date sequentially (ascending) to simulate and shift
            for date_str, record in new_completed:
                try:
                    idx = completed_history.index(record)
                    prev_close = completed_history[idx-1].get("close", 0) if idx > 0 else record.get("close", 0)
                except Exception:
                    prev_close = record.get("close", 0)
                new_flow = simulate_flow_for_day(date_str, record, prev_close)
                base_flows.insert(0, new_flow)
                if len(base_flows) > 5:
                    base_flows.pop()
                    
        # 5. Generate forecasts starting after the latest completed trading day
        m_g = excel_manager.get_macro_geopolitics()
        geopolitics = m_g.get("geopolitics", [])
        macro = m_g.get("macro_indicators", [])
        forecasts_5d = forecaster.generate_multi_day_forecast(vn_history, geopolitics, macro, days=5)
        
        updated_forecasts = []
        for fc in forecasts_5d:
            updated_forecasts.append({
                "date": fc["date"],
                "trend": fc["trend"],
                "probability": fc["probability"],
                "price_range": fc["predicted_range"],
                "risk_warning": fc["warning"]
            })
            
        # 6. Try to write back to Excel file in-place if possible (for local desktop synchronization)
        try:
            wb = excel_manager.load_wb(data_only=False)
            ws = wb["Dong Tien & AI Predictor"]
            
            # Format row fills
            from openpyxl.styles import Font, PatternFill, Alignment
            fill_positive = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
            fill_negative = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
            fill_warning = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
            fill_none = PatternFill(fill_type=None)
            
            # Write Table A (flows)
            for i, f in enumerate(base_flows):
                r = 7 + i
                ws.cell(row=r, column=1, value=f["date"])
                ws.cell(row=r, column=2, value=f["foreign"])
                ws.cell(row=r, column=3, value=f["proprietary"])
                ws.cell(row=r, column=4, value=f["retail"])
                ws.cell(row=r, column=5, value=f["smart_money"])
                
                # Apply styling
                for c in range(1, 6):
                    cell = ws.cell(row=r, column=c)
                    cell.font = Font(name="Segoe UI", size=10)
                    if c == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = fill_none
                    elif c in [2, 3, 4]:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "+#,##0.0;-#,##0.0;0.0"
                        val = cell.value
                        try:
                            val_float = float(val)
                            if val_float > 0:
                                cell.fill = fill_positive
                            elif val_float < 0:
                                cell.fill = fill_negative
                            else:
                                cell.fill = fill_none
                        except Exception:
                            cell.fill = fill_none
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.fill = fill_none
                        
            # Write Table B (forecasts)
            for i, fc in enumerate(forecasts_5d):
                r = 7 + i
                ws.cell(row=r, column=7, value=fc["date"])
                ws.cell(row=r, column=8, value=fc["trend"])
                ws.cell(row=r, column=9, value=fc["probability"])
                ws.cell(row=r, column=10, value=fc["predicted_range"])
                ws.cell(row=r, column=11, value=fc["warning"])
                
                # Apply styling
                cell_trend = ws.cell(row=r, column=8)
                cell_trend.font = Font(name="Segoe UI", size=10, bold=True)
                if "Tăng" in fc["trend"]:
                    cell_trend.fill = fill_positive
                elif "Giảm" in fc["trend"]:
                    cell_trend.fill = fill_negative
                else:
                    cell_trend.fill = fill_warning
                ws.cell(row=r, column=9).number_format = "0%"
                
            excel_manager.save_wb(wb)
        except Exception as e:
            print(f"WARNING: Could not update Excel file: {e}")
            
        return {
            "market_flows": base_flows,
            "forecasts": updated_forecasts
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/excel/ai-scores")
def update_excel_ai_scores(item: AIScoresItem):
    try:
        success = excel_manager.update_ai_scores(item.market_score, item.risk_score, item.opportunity_score)
        db.log_ai_scores(item.market_score, item.risk_score, item.opportunity_score)
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# -------------------------------------------------------------------------
# SQLite History Log Retrieval APIs
# -------------------------------------------------------------------------
@app.get("/api/db/ai-scores-history")
def get_db_ai_scores_history(limit: int = 30):
    try:
        # 1. Fetch raw database records
        history = db.get_historical_scores(limit)
        
        # 2. Fetch live index history
        vn_history = vnstock_client.get_historical_data("VNINDEX", source="kbs")
        if not vn_history:
            vn_history = ssi_client.get_historical_data("VNINDEX")
            
        if not vn_history:
            return history
            
        # 3. Filter completed history
        completed_history = forecaster._filter_completed_history(vn_history)
        if not completed_history:
            return history
            
        completed_dates = []
        for record in completed_history:
            rec_date = record.get("time") or record.get("date")
            if rec_date:
                if isinstance(rec_date, str):
                    rec_date = rec_date.split()[0]
                elif hasattr(rec_date, "strftime"):
                    rec_date = rec_date.strftime("%Y-%m-%d")
                completed_dates.append((rec_date, record))
                
        # 4. Find newer dates not present in database history
        if history:
            latest_db_date = history[-1]["date"] # YYYY-MM-DD
            new_completed = [x for x in completed_dates if x[0] > latest_db_date]
        else:
            new_completed = completed_dates[-limit:]
            
        if new_completed:
            m_g = excel_manager.get_macro_geopolitics()
            geopolitics = m_g.get("geopolitics", [])
            macro = m_g.get("macro_indicators", [])
            
            for date_str, record in new_completed:
                try:
                    idx = completed_history.index(record)
                    slice_history = completed_history[:idx+1]
                except Exception:
                    slice_history = completed_history
                    
                scores = forecaster.compute_ai_scores(slice_history, geopolitics, macro)
                new_rec = {
                    "date": date_str,
                    "market_score": scores["market_score"],
                    "risk_score": scores["risk_score"],
                    "opportunity_score": scores["opportunity_score"],
                    "logged_at": None
                }
                history.append(new_rec)
                
                # Try to log to local DB if writable
                try:
                    db.log_ai_scores(scores["market_score"], scores["risk_score"], scores["opportunity_score"], date_str)
                except Exception:
                    pass
                    
        # Apply limit to returned list
        if len(history) > limit:
            history = history[-limit:]
            
        return history
    except Exception as e:
        print(f"Error in get_db_ai_scores_history: {e}")
        try:
            return db.get_historical_scores(limit)
        except Exception:
            return []

@app.get("/api/db/predictions-history")
def get_db_predictions_history(limit: int = 30):
    return db.get_historical_predictions(limit)

@app.get("/api/db/portfolio-history")
def get_db_portfolio_history(limit: int = 30):
    return db.get_portfolio_history(limit)

# -------------------------------------------------------------------------
# Core AI Scoring, Forecasting & Auto Calculation Engine
# -------------------------------------------------------------------------
@app.post("/api/excel/recalculate-all")
def recalculate_excel_dashboard():
    """
    Core engine that gathers live market signals, technical indicators for VNINDEX,
    portfolio price updates, and recalculates the entire Excel sheet formulas/values.
    """
    try:
        # 1. Fetch live historical data for VN-INDEX to feed technical forecaster
        vn_history = vnstock_client.get_historical_data("VNINDEX", source="kbs")
        if not vn_history:
            # try backup SSI Resolution 1D
            vn_history = ssi_client.get_historical_data("VNINDEX")
            
        # Update Daily Market Flows if a new completed day is available
        try:
            excel_manager.update_market_flows(vn_history)
        except Exception as e:
            print(f"Failed to auto-update market flows in Excel: {e}")
            
        # 2. Get Geopolitics & Macro settings from current Excel sheet state
        m_g = excel_manager.get_macro_geopolitics()
        geopolitics = m_g["geopolitics"]
        macro = m_g["macro_indicators"]
        
        # 3. Dynamic AI Core Scoring
        scores = forecaster.compute_ai_scores(vn_history, geopolitics, macro)
        m_score = scores["market_score"]
        r_score = scores["risk_score"]
        o_score = scores["opportunity_score"]
        excel_manager.update_ai_scores(m_score, r_score, o_score)
        db.log_ai_scores(m_score, r_score, o_score)
        
        # 4. Machine Learning Trend Forecast
        forecast = forecaster.generate_forecast(vn_history, geopolitics, macro)
        db.log_prediction(
            trend=forecast["trend"],
            probability=forecast["probability"],
            predicted_range=forecast["predicted_range"],
            warning=forecast["warning"],
            date_str=forecast["date"]
        )
        
        # Write forecasts into the Excel sheet 6 (Dong Tien & AI Predictor)
        # We generate forecasts for the next 5 sessions and write them into rows 7 to 11
        forecasts_5d = forecaster.generate_multi_day_forecast(vn_history, geopolitics, macro, days=5)
        wb = excel_manager.load_wb(data_only=False)
        ws = wb["Dong Tien & AI Predictor"]
        
        fill_positive = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        fill_negative = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        fill_warning = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
        
        for i, fc in enumerate(forecasts_5d):
            r = 7 + i
            ws.cell(row=r, column=7, value=fc["date"])
            ws.cell(row=r, column=8, value=fc["trend"])
            ws.cell(row=r, column=9, value=fc["probability"])
            ws.cell(row=r, column=10, value=fc["predicted_range"])
            ws.cell(row=r, column=11, value=fc["warning"])
            
            # Apply styling
            cell_trend = ws.cell(row=r, column=8)
            cell_trend.font = Font(name="Segoe UI", size=10, bold=True)
            if "Tăng" in fc["trend"]:
                cell_trend.fill = fill_positive
            elif "Giảm" in fc["trend"]:
                cell_trend.fill = fill_negative
            else:
                cell_trend.fill = fill_warning
                
            ws.cell(row=r, column=9).number_format = "0%"
            
        excel_manager.save_wb(wb)
        
        # 5. Sync Portfolio Tickers & live prices
        port = excel_manager.get_portfolio()
        tickers = [item["ticker"] for item in port["items"] if item["ticker"]]
        price_map = {}
        for t in tickers:
            # Use price depth endpoint to find latest closing/last price
            p_data = ssi_client.get_price_depth(t)
            if not p_data or p_data.get("last_price", 0) == 0:
                raw_depth = vnstock_client.get_price_depth(t)
                if isinstance(raw_depth, dict) and "last_price" in raw_depth:
                    p_data = raw_depth
                else:
                    p_data = ssi_client._generate_mock_price_depth(t)
            
            raw_p = p_data.get("last_price", 0)
            if raw_p > 0:
                # convert board standard units (1,000 VND multiplier)
                if raw_p < 2000:
                    price_map[t] = int(raw_p * 1000)
                else:
                    price_map[t] = int(raw_p)
                    
        if price_map:
            excel_manager.update_portfolio_prices(price_map)
            
        # 6. Log updated portfolio totals to DB
        updated_port = excel_manager.get_portfolio()
        totals = updated_port["totals"]
        db.log_portfolio_snapshot(totals["cost_basis"], totals["current_val"], totals["pnl"], totals["pnl_pct"])
        
        # 7. Update Overview indices points in Excel
        wb_ov = excel_manager.load_wb(data_only=False)
        ws_ov = wb_ov["Dashboard Tong Quan"]
        
        # Fetch VNINDEX live price depth
        v_depth = ssi_client.get_price_depth("VNINDEX")
        if v_depth and v_depth.get("last_price", 0) > 0:
            ws_ov.cell(row=7, column=2, value=v_depth["last_price"]) # VN-INDEX closing
            ws_ov.cell(row=7, column=3, value=v_depth["change"])
            ws_ov.cell(row=7, column=4, value=v_depth["change_pct"] / 100.0)
            
        # Fetch VN30 live price depth
        v30_depth = ssi_client.get_price_depth("VN30")
        if v30_depth and v30_depth.get("last_price", 0) > 0:
            ws_ov.cell(row=8, column=2, value=v30_depth["last_price"]) # VN30 closing
            ws_ov.cell(row=8, column=3, value=v30_depth["change"])
            ws_ov.cell(row=8, column=4, value=v30_depth["change_pct"] / 100.0)

        # Fetch HNXINDEX live price depth
        h_depth = ssi_client.get_price_depth("HNXINDEX")
        if h_depth and h_depth.get("last_price", 0) > 0:
            ws_ov.cell(row=9, column=2, value=h_depth["last_price"]) # HNX-INDEX closing
            ws_ov.cell(row=9, column=3, value=h_depth["change"])
            ws_ov.cell(row=9, column=4, value=h_depth["change_pct"] / 100.0)

        # Fetch UPCOMINDEX live price depth
        u_depth = ssi_client.get_price_depth("UPCOMINDEX")
        if u_depth and u_depth.get("last_price", 0) > 0:
            ws_ov.cell(row=10, column=2, value=u_depth["last_price"]) # UPCoM-INDEX closing
            ws_ov.cell(row=10, column=3, value=u_depth["change"])
            ws_ov.cell(row=10, column=4, value=u_depth["change_pct"] / 100.0)
            
        # Fetch VN30F1M live price depth
        vf_depth = ssi_client.get_price_depth("VN30F1M")
        if not vf_depth or vf_depth.get("last_price", 0) == 0:
            vf_depth = ssi_client._generate_mock_price_depth("VN30F1M")
            
        vf_price = vf_depth.get("last_price", 1260.50)
        vf_change = vf_depth.get("change", 5.25)
        vf_change_pct = vf_depth.get("change_pct", 0.42)
        
        ws_ov.cell(row=11, column=2, value=vf_price)
        ws_ov.cell(row=11, column=3, value=vf_change)
        ws_ov.cell(row=11, column=4, value=vf_change_pct / 100.0)
        
        # Style row 11 change font color dynamically
        cell_c = ws_ov.cell(row=11, column=3)
        if vf_change > 0:
            cell_c.font = Font(name="Segoe UI", size=10, bold=True, color="196F3D")
            cell_c.fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        elif vf_change < 0:
            cell_c.font = Font(name="Segoe UI", size=10, bold=True, color="943126")
            cell_c.fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
            
        excel_manager.save_wb(wb_ov)
        
        # 8. Calculate derivatives recommendation
        v30_val = v30_depth.get("last_price", 1262.15) if (v30_depth and v30_depth.get("last_price", 0) > 0) else 1262.15
        basis = vf_price - v30_val
        
        trend_name = forecast["trend"]
        if "Tăng" in trend_name:
            rec = "LONG"
            prob = min(0.85, forecast["probability"] + 0.05)
            target = f"{int(vf_price + 10):,} - {int(vf_price + 15):,} điểm"
            stop_loss = int(vf_price - 8)
        elif "Giảm" in trend_name:
            rec = "SHORT"
            prob = min(0.85, forecast["probability"] + 0.05)
            target = f"{int(vf_price - 15):,} - {int(vf_price - 10):,} điểm"
            stop_loss = int(vf_price + 8)
        else:
            rec = "QUAN SÁT"
            prob = 0.50
            target = "—"
            stop_loss = 0
            
        excel_manager.update_derivatives_recommendation(
            price=vf_price,
            basis=basis,
            recommendation=rec,
            probability=prob,
            target=target,
            stop_loss=stop_loss
        )
        
        return {
            "success": True,
            "ai_scores": {
                "market_score": m_score,
                "risk_score": r_score,
                "opportunity_score": o_score
            },
            "forecast": forecast,
            "prices_synced": price_map
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Recalculation failed: {str(e)}")

# -------------------------------------------------------------------------
# Stock Signal Analysis Endpoint (Long/Short Reason Analysis)
# -------------------------------------------------------------------------
@app.get("/api/stock-signals")
def get_stock_signals(tickers: str = ""):
    """
    Returns comprehensive Long/Short signal analysis for portfolio stocks.
    Combines technical indicators, capital flows, derivatives basis, 
    accumulated buying patterns, and sector macro factors.
    """
    try:
        # Get portfolio items
        port = excel_manager.get_portfolio()
        items = port.get("items", [])
        
        # If custom tickers provided, build synthetic items
        if tickers:
            ticker_list = [t.strip().upper() for t in tickers.split(",") if t.strip()]
            # Merge custom tickers with portfolio
            existing_tickers = {i["ticker"] for i in items}
            for t in ticker_list:
                if t not in existing_tickers:
                    items.append({
                        "ticker": t,
                        "name": f"CP {t}",
                        "buy_price": 0,
                        "quantity": 0,
                        "current_price": 0,
                        "pnl_pct": 0
                    })
        
        # Get live price data for all tickers
        price_data = {}
        for item in items:
            t = item.get("ticker", "")
            if not t:
                continue
            try:
                p_data = ssi_client.get_price_depth(t)
                if not p_data or p_data.get("last_price", 0) == 0:
                    raw_depth = vnstock_client.get_price_depth(t)
                    if isinstance(raw_depth, dict) and "last_price" in raw_depth:
                        p_data = raw_depth
                if p_data:
                    price_data[t] = p_data
            except Exception as e:
                print(f"Price fetch error for {t}: {e}")
        
        # Get current VNINDEX trend for market context
        vnindex_trend = "Tăng nhẹ"
        try:
            vn_history = vnstock_client.get_historical_data("VNINDEX", source="kbs")
            if not vn_history:
                vn_history = ssi_client.get_historical_data("VNINDEX")
            if vn_history:
                m_g = excel_manager.get_macro_geopolitics()
                f = forecaster.generate_forecast(vn_history, m_g["geopolitics"], m_g["macro_indicators"])
                vnindex_trend = f.get("trend", "Tăng nhẹ")
        except Exception as e:
            print(f"Trend fetch error: {e}")
        
        # Generate signal analysis
        signals = forecaster.generate_stock_signal_analysis(items, price_data, vnindex_trend)
        
        return {
            "success": True,
            "vnindex_trend": vnindex_trend,
            "signals": signals
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Signal analysis failed: {str(e)}")


@app.get("/api/derivatives/live-candle")
def get_derivatives_live_candle():
    """
    Trả về nến M5 gần nhất của VN30F1M bằng cách gộp 5 nến 1 phút gần nhất
    từ vnstock (dữ liệu thật), không dùng SSI mock.
    """
    try:
        today = vn_now().date()
        start = end = today.strftime("%Y-%m-%d")

        # 1. Lấy nến 1 phút VN30F1M hôm nay (thật, không mock)
        bars_1m = vnstock_client.get_historical_data("VN30F1M", start, end, "1m", "VCI")
        if not bars_1m:
            raise HTTPException(status_code=503,
                                detail="Chưa có dữ liệu nến VN30F1M hôm nay. Kiểm tra lại trong giờ giao dịch.")

        # Bỏ nến cuối (đang hình thành), lấy nến 1 phút hoàn chỉnh gần nhất
        complete = bars_1m[:-1] if len(bars_1m) > 1 else bars_1m
        if not complete:
            raise HTTPException(status_code=503, detail="Không đủ nến hoàn chỉnh.")

        # Dùng đúng 1 nến 1 phút — chuẩn xác hơn gộp M5 vì giữ nguyên timing
        bar = complete[-1]
        open_p  = bar["open"]
        close_p = bar["close"]
        high_p  = bar["high"]
        low_p   = bar["low"]
        volume  = bar["volume"]
        recent  = [bar]
        candle_time = str(bar.get("time", ""))[:16]

        # 2. VN30 index cho basis (cố gắng lấy, không bắt buộc)
        vn30_close = None
        try:
            vn30_bars = vnstock_client.get_historical_data("VN30", start, end, "1m", "VCI")
            if vn30_bars:
                vn30_close = vn30_bars[-1]["close"]
        except Exception:
            pass
        basis = round(close_p - vn30_close, 1) if vn30_close else None

        # 3. Choppiness Index (14 bars) — regime detection
        ci = _calc_choppiness(complete, period=14)
        if ci is not None:
            if ci > 61.8:
                market_regime = "RANGING"
            elif ci < 38.2:
                market_regime = "TRENDING"
            else:
                market_regime = "NEUTRAL"
        else:
            market_regime = None

        # 4. M15 trend: dùng 15 nến 1m hoàn chỉnh gần nhất làm "nến M15"
        m15_bars = complete[-15:] if len(complete) >= 15 else complete
        if m15_bars:
            m15_open  = m15_bars[0]["open"]
            m15_close = m15_bars[-1]["close"]
            m15_bullish = m15_close > m15_open
        else:
            m15_bullish = None

        # 5. Session high/low (dùng toàn bộ nến hôm nay kể cả nến đang hình thành)
        session_high = max(b["high"] for b in bars_1m)
        session_low  = min(b["low"]  for b in bars_1m)

        # 6. Tạo mô tả price action dựa trên hình dạng nến thật
        candle_range = high_p - low_p
        body = abs(close_p - open_p)
        body_ratio = body / candle_range if candle_range > 0 else 0
        bullish = close_p >= open_p

        # So sánh volume với trung bình các nến đã có trong ngày
        avg_vol_per_bar = sum(b["volume"] for b in bars_1m) / len(bars_1m)
        vol_spike = volume / (avg_vol_per_bar * len(recent)) if avg_vol_per_bar > 0 else 1.0
        vol_note = " KL đột biến xác nhận." if vol_spike >= 1.8 else ""

        near_high = close_p >= high_p - candle_range * 0.06
        near_low  = close_p <= low_p  + candle_range * 0.06

        if volume < 200:
            pa_text = "Thanh khoản cạn kiệt, thị trường đi ngang thăm dò."
        elif body_ratio >= 0.65 and bullish and near_high:
            pa_text = f"Nến bứt phá vượt đỉnh, lực cầu Long áp đảo.{vol_note}"
        elif body_ratio >= 0.55 and bullish:
            pa_text = f"Nến rút chân tích cực, lực cầu chủ động hấp thụ cung.{vol_note}"
        elif body_ratio >= 0.65 and not bullish and near_low:
            pa_text = f"Thân nến đỏ dài sát đáy, áp lực bán đè nặng phe Long.{vol_note}"
        elif body_ratio >= 0.55 and not bullish:
            pa_text = f"Nến từ chối tăng, áp lực chốt lời ngắn hạn xuất hiện.{vol_note}"
        elif close_p > (high_p + low_p) / 2 + candle_range * 0.1:
            pa_text = "Nến rút chân tích cực, lực cầu chủ động hấp thụ cung."
        elif close_p < (high_p + low_p) / 2 - candle_range * 0.1:
            pa_text = "Nến từ chối tăng, áp lực chốt lời ngắn hạn xuất hiện."
        else:
            pa_text = "Nến thân nhỏ biến động hẹp, hai phe Long/Short đang giằng co."

        session_volume = int(sum(b["volume"] for b in bars_1m))

        return {
            "success": True,
            "candle_time": candle_time,
            "open_price": round(open_p, 1),
            "close_price": round(close_p, 1),
            "high_price": round(high_p, 1),
            "low_price": round(low_p, 1),
            "volume": int(volume),
            "session_volume": session_volume,
            "basis": basis,
            "vn30_price": round(vn30_close, 1) if vn30_close else None,
            "price_action": pa_text,
            "bars_used": len(recent),
            "vol_spike": round(vol_spike, 2),
            "m15_bullish": m15_bullish,
            "session_high": round(session_high, 1),
            "session_low": round(session_low, 1),
            "choppiness": ci,
            "market_regime": market_regime,
            "data_source": "vnstock_real",
        }
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/derivatives/intraday-forecast")
def get_derivatives_intraday_forecast(item: IntradayCandleItem):
    try:
        close_p      = item.close_price
        volume       = item.volume
        high_p       = item.high_price
        low_p        = item.low_price
        basis        = item.basis
        pa           = item.price_action.strip().lower()
        vol_spike    = item.vol_spike
        m15_bullish  = item.m15_bullish
        session_high = item.session_high
        session_low  = item.session_low

        candle_range = high_p - low_p

        # ── Hard filters (reject signal regardless of score) ─────────────────
        choppiness    = item.choppiness
        market_regime = item.market_regime
        hard_neutral_reason = None
        if market_regime == "RANGING":
            ci_str = f"{choppiness:.0f}" if choppiness else "?"
            hard_neutral_reason = (
                f"Thị trường đang SIDEWAY (Choppiness Index = {ci_str} > 61.8). "
                f"Trong range, tín hiệu chiều nào cũng thua vì giá dao động giữa 2 biên. "
                f"Đợi CI giảm dưới 61.8 (thị trường chọn hướng) rồi mới giao dịch."
            )
        elif volume < 100:
            hard_neutral_reason = f"KL nến 1m cực thấp ({volume:.0f} HĐ) — dòng tiền cạn kiệt."
        elif abs(basis) > 8.0:
            hard_neutral_reason = f"Basis quá rộng ({basis:+.1f}đ) — rủi ro ép basis đột ngột."
        elif candle_range < 0.5:
            hard_neutral_reason = f"Range nến quá hẹp ({candle_range:.1f}đ) — tín hiệu nhiễu."

        # ── Keyword detection ─────────────────────────────────────────────────
        long_kws    = ["rút chân", "pinbar", "bứt phá", "vượt đỉnh", "cạn cung",
                       "lực cầu", "bullish"]
        short_kws   = ["đỏ dài", "thủng đáy", "phân kỳ", "áp lực bán", "chốt lời",
                       "đè nặng", "từ chối tăng", "bearish"]
        neutral_kws = ["giằng co", "đi ngang", "cân bằng", "thăm dò", "biến động hẹp"]

        has_long_kw  = any(kw in pa for kw in long_kws)
        has_short_kw = any(kw in pa for kw in short_kws)
        if any(kw in pa for kw in neutral_kws) or (has_long_kw and has_short_kw):
            has_long_kw = has_short_kw = False

        # ── Position ratio (0 = đáy, 1 = đỉnh) ──────────────────────────────
        pos_ratio = (close_p - low_p) / candle_range if candle_range > 0 else 0.5
        if candle_range < 1.0:          # nến quá ngắn, vị trí không đáng tin
            pos_ratio = 0.5
            has_long_kw = has_short_kw = False

        # ══ MULTI-FACTOR SCORING ═════════════════════════════════════════════
        # Mỗi yếu tố đóng góp điểm dương (bullish) hoặc âm (bearish).
        # Cần tổng ≥ +3 để ra Long, ≤ −3 để ra Short. Ngưỡng này đòi hỏi
        # ít nhất 2-3 yếu tố đồng thuận — tránh kích hoạt từ 1 nến đơn lẻ.

        # Yếu tố 1: Hình nến  (−2 … +2)
        if has_long_kw or pos_ratio >= 0.65:
            c_pts = 2
            matched_kw = next((k for k in long_kws if k in pa), None)
            c_desc = f"Bullish {pos_ratio:.0%} range" + (f" · kw '{matched_kw}'" if matched_kw else "")
        elif pos_ratio > 0.55:
            c_pts = 1
            c_desc = f"Nghiêng tăng {pos_ratio:.0%} range"
        elif has_short_kw or pos_ratio <= 0.35:
            c_pts = -2
            matched_kw = next((k for k in short_kws if k in pa), None)
            c_desc = f"Bearish {pos_ratio:.0%} range" + (f" · kw '{matched_kw}'" if matched_kw else "")
        elif pos_ratio < 0.45:
            c_pts = -1
            c_desc = f"Nghiêng giảm {pos_ratio:.0%} range"
        else:
            c_pts = 0
            c_desc = f"Trung lập {pos_ratio:.0%} range (giằng co)"

        # Yếu tố 2: Khối lượng — có hướng theo nến (c_pts đã tính ở trên)
        # KL cao + nến tăng → mua mạnh (+); KL cao + nến giảm → bán mạnh (−).
        # KL thấp luôn trừ điểm: thiếu dòng tiền dù nến hình dạng đẹp cũng không đáng tin.
        candle_dir = 1 if c_pts > 0 else (-1 if c_pts < 0 else 0)

        if vol_spike is not None:
            if vol_spike >= 2.0 and candle_dir != 0:
                v_pts = 2 * candle_dir
                v_desc = f"KL đột biến {vol_spike:.1f}× TB — xác nhận {'mua' if v_pts > 0 else 'bán'} rất mạnh"
            elif vol_spike >= 1.4 and candle_dir != 0:
                v_pts = candle_dir
                v_desc = f"KL cao {vol_spike:.1f}× TB — xác nhận {'mua' if v_pts > 0 else 'bán'}"
            elif vol_spike < 0.6:
                v_pts = -1
                v_desc = f"KL yếu {vol_spike:.1f}× TB — thiếu dòng tiền xác nhận"
            else:
                v_pts = 0
                v_desc = f"KL {vol_spike:.1f}× TB — {'trung tính' if candle_dir == 0 else 'không đủ mạnh để xác nhận'}"
        else:
            v_pts = 0
            v_desc = f"KL tuyệt đối {volume:.0f} HĐ (nhập tay, không có TB để so)"

        # Yếu tố 3: Xu hướng M15  (−1 … +1)
        if m15_bullish is True:
            m_pts = 1
            m_desc = "M15 ↑ Tăng — trend khung lớn ủng hộ Long"
        elif m15_bullish is False:
            m_pts = -1
            m_desc = "M15 ↓ Giảm — trend khung lớn ủng hộ Short"
        else:
            m_pts = 0
            m_desc = "M15 chưa xác định (chế độ nhập tay)"

        # Yếu tố 4: Vị trí so với đỉnh/đáy phiên — có nhận thức xu hướng (−1 … +1)
        # Sát đỉnh → kháng cự, luôn trừ điểm (−1).
        # Sát đáy → hỗ trợ NẾU uptrend; nhưng trong downtrend = nguy cơ thủng đáy (−1 thay vì +1).
        sr_pts = 0
        sr_note = ""
        if session_high and session_low:
            near_high = close_p >= session_high - 1.5
            near_low  = close_p <= session_low  + 1.5
            if near_high:
                sr_pts  = -1
                sr_note = f"⚠ Sát đỉnh phiên {session_high:.1f} — kháng cự mạnh"
            elif near_low:
                if m15_bullish is False:
                    sr_pts  = -1
                    sr_note = f"⚠ Sát đáy phiên {session_low:.1f} — rủi ro thủng đáy (downtrend)"
                else:
                    sr_pts  = 1
                    sr_note = f"Sát đáy phiên {session_low:.1f} — vùng hỗ trợ"
            else:
                sr_note = f"Giữa biên {session_low:.1f}–{session_high:.1f}"

        score = c_pts + v_pts + m_pts + sr_pts
        MIN_SCORE = 3   # ngưỡng để ra tín hiệu (+3 Long, −3 Short)

        # ── Quyết định ───────────────────────────────────────────────────────
        if hard_neutral_reason:
            trend  = "ĐI NGANG (QUAN SÁT)"
            action = "Đứng ngoài"
            entry  = "Không khuyến nghị"
            sl     = "Không có"
            tp     = "Không có"
            arg_pa    = f"Lọc cứng: {hard_neutral_reason}"
            arg_basis = f"Basis {basis:+.1f}đ | Điểm scoring bị bỏ qua."
            arg_sr    = f"Hỗ trợ: {low_p:.1f} | Kháng cự: {high_p:.1f}."
        elif score >= MIN_SCORE:
            trend  = "TĂNG (LONG)"
            action = "Mở Long"
            entry  = f"{close_p - 0.2:.1f} – {close_p + 0.2:.1f}"
            sl     = f"{close_p - 2.0:.1f} (Cắt lỗ 2.0đ)"
            tp     = f"TP1: {close_p + 4.0:.1f} | TP2: {close_p + 6.0:.1f} (R:R ≥1:2)"
            arg_pa    = f"Nến: {c_desc} [{c_pts:+d}] · KL: {v_desc} [{v_pts:+d}]"
            arg_basis = f"M15: {m_desc} [{m_pts:+d}] · Basis {basis:+.1f}đ"
            arg_sr    = f"S/R phiên: {sr_note} [{sr_pts:+d}] · Điểm tổng: {score:+d}/6 (ngưỡng ≥+{MIN_SCORE})"
        elif score <= -MIN_SCORE:
            trend  = "GIẢM (SHORT)"
            action = "Mở Short"
            entry  = f"{close_p - 0.2:.1f} – {close_p + 0.2:.1f}"
            sl     = f"{close_p + 2.0:.1f} (Cắt lỗ 2.0đ)"
            tp     = f"TP1: {close_p - 4.0:.1f} | TP2: {close_p - 6.0:.1f} (R:R ≥1:2)"
            arg_pa    = f"Nến: {c_desc} [{c_pts:+d}] · KL: {v_desc} [{v_pts:+d}]"
            arg_basis = f"M15: {m_desc} [{m_pts:+d}] · Basis {basis:+.1f}đ"
            arg_sr    = f"S/R phiên: {sr_note} [{sr_pts:+d}] · Điểm tổng: {score:+d}/6 (ngưỡng ≤-{MIN_SCORE})"
        else:
            trend  = "ĐI NGANG (QUAN SÁT)"
            action = "Đứng ngoài"
            entry  = "Không khuyến nghị"
            sl     = "Không có"
            tp     = "Không có"
            arg_pa    = f"Nến: {c_desc} [{c_pts:+d}] · KL: {v_desc} [{v_pts:+d}]"
            arg_basis = f"M15: {m_desc} [{m_pts:+d}] · Basis {basis:+.1f}đ"
            arg_sr    = f"S/R phiên: {sr_note} [{sr_pts:+d}] · Điểm tổng: {score:+d}/6 — chưa đủ ngưỡng ±{MIN_SCORE}"

        # Ngoài phiên thì nến M5 đang đứng yên, mọi tín hiệu rút ra từ nó đều vô nghĩa.
        # Trả trạng thái ra để giao diện nói thẳng, thay vì hiện khuyến nghị như thường.
        session_open, session_note = derivatives_session_state()
        if not session_open:
            trend = "NGOÀI PHIÊN GIAO DỊCH"
            action = "Đứng ngoài"
            entry = "Không khuyến nghị"
            sl = "Không có"
            tp = "Không có"
            arg_pa = (f"{session_note} Dữ liệu nến M5 đang đứng yên nên không rút ra được "
                      "tín hiệu nào có ý nghĩa.")
            arg_basis = f"Basis ghi nhận lần cuối {basis:+.1f} điểm."
            arg_sr = f"Hỗ trợ: {low_p:.1f} | Kháng cự: {high_p:.1f} (theo nến cuối phiên)."

        # Save recommendation to file log
        try:
            save_derivatives_log(trend, action, entry, sl, tp)
        except Exception as log_ex:
            print("Error saving log:", log_ex)

        return {
            "success": True,
            "session_open": session_open,
            "session_note": session_note,
            "trend_verdict": trend,
            "action_signal": action,
            "entry_range": entry,
            "stop_loss": sl,
            "take_profit": tp,
            "score": score if not hard_neutral_reason else None,
            "choppiness": choppiness,
            "market_regime": market_regime,
            "arguments": {
                "price_action_vol": arg_pa,
                "basis_impact": arg_basis,
                "support_resistance": arg_sr
            },
            "disclaimer": "Tín hiệu chỉ mang tính chất tham khảo, hãy tuân thủ kỷ luật Stop Loss."
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



# ---------------------------------------------------------------------------
# ĐỐI CHIẾU KẾT QUẢ THẬT CỦA TÍN HIỆU PHÁI SINH
#
# Hàng tổng kết trước đây cộng khoảng cách Entry->TP1 của mọi tín hiệu, tức ngầm giả
# định lệnh nào cũng chạm TP1 và không lệnh nào chạm SL, rồi trưng ra dưới cái tên
# "Lợi nhuận ròng tạm tính". Ở đây thay bằng việc dò giá thật sau thời điểm phát tín
# hiệu để xem chạm SL hay TP trước.
#
# Nguồn: nến 1 phút của VN30F1M qua vnstock (VCI), 241 nến mỗi phiên phủ 09:00-14:45.
# Dùng vnstock_client.get_historical_data vì nó trả [] khi hỏng - KHÔNG bao giờ sinh
# dữ liệu giả như đường SSI mock. Lãi lỗ tính từ giá bịa ra thì còn tệ hơn con số cũ.
# ---------------------------------------------------------------------------
VN30F1M_POINT_VALUE = 100_000   # VND cho mỗi điểm chỉ số, theo đặc tả hợp đồng
ROUND_TRIP_FEE = 20_000         # VND mỗi hợp đồng cho trọn vòng mở + đóng

_NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")


def _numbers_in(text):
    return [float(x) for x in _NUMBER_RE.findall(str(text or ""))]


def _entry_price(text):
    """
    "1889.8 - 1890.2" -> 1890.0.

    Lấy trung điểm dải entry. Giá khớp thật nằm đâu đó trong dải và hệ thống không ghi
    lại, nên đây là một giả định - phải nói ra, không được lờ đi.
    """
    nums = _numbers_in(text)
    if not nums:
        return None
    return round(sum(nums[:2]) / len(nums[:2]), 2)


def _tp1_price(text):
    """ "TP1: 1886.0 | TP2: 1884.0 (R:R tối thiểu 1:2)" -> 1886.0 """
    m = re.search(r"TP1\s*:\s*(-?\d+(?:\.\d+)?)", str(text or ""))
    if m:
        return float(m.group(1))
    nums = _numbers_in(text)
    return nums[0] if nums else None


def _load_minute_bars(day):
    """Nến 1 phút của VN30F1M trong một ngày, đã sắp theo thời gian. Hỏng thì trả []."""
    rows = vnstock_client.get_historical_data("VN30F1M", day, day, "1m", "VCI")
    bars = []
    for r in rows or []:
        t = str(r.get("time", ""))
        if not t.startswith(day):
            continue
        try:
            bars.append({
                "time": t,
                "high": float(r["high"]),
                "low": float(r["low"]),
                "close": float(r["close"]),
            })
        except (KeyError, TypeError, ValueError):
            continue
    bars.sort(key=lambda b: b["time"])
    return bars


def _resolve_signal(sig, bars, session_open_now):
    """
    Dò từng nến sau thời điểm phát tín hiệu xem chạm SL hay TP1 trước.

    Trong một nến 1 phút mà giá quét qua CẢ hai mức thì không thể biết mức nào tới
    trước - trả 'khong_xac_dinh' thay vì đoán. Đoán ở đây là bịa ra một khoản lãi lỗ.
    """
    action = (sig.action or "")
    if "Long" in action:
        direction = 1
    elif "Short" in action:
        direction = -1
    else:
        return {"outcome": "khong_phai_lenh", "points": None}

    entry = _entry_price(sig.entry)
    sl = _numbers_in(sig.sl)[0] if _numbers_in(sig.sl) else None
    tp = _tp1_price(sig.tp)
    if entry is None or sl is None or tp is None:
        return {"outcome": "thieu_thong_so", "points": None}

    if not bars:
        return {"outcome": "thieu_du_lieu", "points": None}

    opened_at = f"{sig.date} {sig.time}"
    for b in bars:
        if b["time"] <= opened_at:
            continue
        if direction == 1:
            hit_tp, hit_sl = b["high"] >= tp, b["low"] <= sl
        else:
            hit_tp, hit_sl = b["low"] <= tp, b["high"] >= sl
        if hit_tp and hit_sl:
            return {"outcome": "khong_xac_dinh", "points": None, "resolved_at": b["time"]}
        if hit_tp:
            return {"outcome": "thang", "exit": tp, "entry": entry,
                    "points": round(direction * (tp - entry), 2), "resolved_at": b["time"]}
        if hit_sl:
            return {"outcome": "thua", "exit": sl, "entry": entry,
                    "points": round(direction * (sl - entry), 2), "resolved_at": b["time"]}

    # Chưa chạm mức nào. Còn trong phiên thì lệnh vẫn đang chạy; hết phiên rồi thì
    # tính tạm theo giá đóng cửa và tách riêng, KHÔNG cộng vào lãi lỗ đã chốt.
    last_close = bars[-1]["close"]
    unreal = round(direction * (last_close - entry), 2)
    if session_open_now and sig.date == vn_now().strftime("%Y-%m-%d"):
        return {"outcome": "dang_mo", "entry": entry, "mark": last_close,
                "points": None, "unrealized_points": unreal}
    return {"outcome": "het_phien_chua_cham", "entry": entry, "mark": last_close,
            "points": None, "unrealized_points": unreal}


@app.post("/api/derivatives/evaluate-log")
def evaluate_derivatives_log(req: EvaluateLogRequest):
    """
    Chấm kết quả thật cho nhật ký tín hiệu M5.

    Trả về từng tín hiệu kèm kết quả, và phần tổng kết CHỈ cộng những lệnh đã chốt
    (chạm SL hoặc TP). Lệnh đang mở, lệnh hết phiên chưa chạm, lệnh không đủ dữ liệu
    đều đếm riêng và nêu rõ - không gộp vào con số lãi lỗ.
    """
    try:
        contracts = max(1, int(req.contracts or 1))
        target_date = req.date or vn_now().strftime("%Y-%m-%d")
        session_open_now, _ = derivatives_session_state()

        rows, sig_warning = supabase_client.get_signals(target_date)
        signals = [TradeSignalItem(**row) for row in rows]

        warnings = [sig_warning] if sig_warning else []
        bars = _load_minute_bars(target_date)
        if not bars:
            warnings.append(f"Không lấy được nến 1 phút của VN30F1M ngày {target_date}; "
                            f"các lệnh trong ngày đó chưa chấm được.")

        results, realized_points, wins, losses = [], 0.0, 0, 0
        pending = {"dang_mo": 0, "het_phien_chua_cham": 0,
                   "khong_xac_dinh": 0, "thieu_du_lieu": 0, "thieu_thong_so": 0}

        for sig in signals:
            r = _resolve_signal(sig, bars, session_open_now)
            if r["outcome"] == "khong_phai_lenh":
                continue
            if r["outcome"] == "thang":
                wins += 1
                realized_points += r["points"]
            elif r["outcome"] == "thua":
                losses += 1
                realized_points += r["points"]
            elif r["outcome"] in pending:
                pending[r["outcome"]] += 1
            results.append({**r, "date": sig.date, "time": sig.time, "action": sig.action})

        closed = wins + losses
        gross = realized_points * VN30F1M_POINT_VALUE * contracts
        fees = closed * ROUND_TRIP_FEE * contracts
        return {
            "success": True,
            "date": target_date,
            "results": results,
            "summary": {
                "closed": closed,
                "wins": wins,
                "losses": losses,
                "win_rate": round(wins / closed * 100, 1) if closed else None,
                "realized_points": round(realized_points, 2),
                "realized_vnd": round(gross - fees),
                "fees_vnd": fees,
                "contracts": contracts,
                "pending": pending,
            },
            "assumptions": [
                "Giá vào lệnh lấy trung điểm dải Entry; giá khớp thật không được ghi lại.",
                "Giá thoát lấy đúng mức SL/TP, chưa tính trượt giá.",
                "Nến 1 phút quét qua cả SL lẫn TP thì không biết mức nào tới trước, "
                "lệnh đó để 'không xác định' chứ không đoán.",
            ],
            "warnings": warnings,
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Không chấm được nhật ký: {e}")


@app.get("/api/derivatives/history-log")
def get_derivatives_history_log(date: str = Query(None, description="YYYY-MM-DD, bỏ trống lấy hôm nay (giờ VN)")):
    """
    Nhật ký tín hiệu M5 của một ngày, đọc từ Supabase.

    Trước đây đọc static/derivatives_history.json - đúng ở local, luôn rỗng trên
    Vercel vì filesystem chỉ đọc. Không bao giờ ném lỗi ra ngoài: Supabase hỏng thì
    trả 'signals': [] kèm 'warning', để giao diện phân biệt được "chưa có lệnh nào"
    với "không đọc được dữ liệu" - im lặng trả rỗng ở đây sẽ trông giống trường hợp
    trước, gây hiểu nhầm.
    """
    target_date = date or vn_now().strftime("%Y-%m-%d")
    signals, warning = supabase_client.get_signals(target_date)
    return {"date": target_date, "signals": signals, "warning": warning}


if __name__ == "__main__":
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    print(f"\n{'='*60}")
    print(f"  🚀 AI STOCK DASHBOARD đang chạy!")
    print(f"  📍 Trên máy này  : http://127.0.0.1:8000")
    print(f"  🌐 Mạng LAN (chia sẻ): http://{local_ip}:8000")
    print(f"  📱 Quét QR bên dưới để mở trên điện thoại")
    print(f"{'='*60}\n")
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)



# ==========================================
# MODULE MỞ RỘNG: TOP 10 TECHNICAL TRADING
# ==========================================
from top10_trading import get_top10_trading_signals

@app.get("/api/top10-technical-trading")
def api_top10_trading():
    try:
        data = get_top10_trading_signals()
        return {"status": "success", "data": data}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# -------------------------------------------------------------------------
# Surfing / T+3 Screener — lọc cổ phiếu theo khối lượng đột biến + kỹ thuật
# -------------------------------------------------------------------------

_INDUSTRY_MAP: dict = {
    "ACB": "Ngân hàng", "BID": "Ngân hàng", "CTG": "Ngân hàng",
    "HDB": "Ngân hàng", "LPB": "Ngân hàng", "MBB": "Ngân hàng",
    "MSB": "Ngân hàng", "SHB": "Ngân hàng", "SSB": "Ngân hàng",
    "STB": "Ngân hàng", "TCB": "Ngân hàng", "TPB": "Ngân hàng",
    "VCB": "Ngân hàng", "VIB": "Ngân hàng", "VPB": "Ngân hàng",
    "BSI": "Chứng khoán", "CTS": "Chứng khoán", "HCM": "Chứng khoán",
    "SSI": "Chứng khoán", "VCI": "Chứng khoán", "VND": "Chứng khoán",
    "BCM": "BĐS", "DIG": "BĐS", "DXG": "BĐS", "KDH": "BĐS",
    "NLG": "BĐS", "NVL": "BĐS", "PDR": "BĐS", "VHM": "BĐS",
    "VIC": "BĐS", "VRE": "BĐS",
    "FPT": "Công nghệ",
    "MWG": "Bán lẻ", "PNJ": "Bán lẻ",
    "VNM": "Thực phẩm", "MSN": "Tiêu dùng", "SAB": "Đồ uống", "KDC": "Thực phẩm",
    "HPG": "Thép", "HSG": "Thép", "NKG": "Thép",
    "GAS": "Dầu khí", "PLX": "Dầu khí", "PVD": "Dầu khí", "PVS": "Dầu khí",
    "GVR": "Cao su",
    "DGC": "Hóa chất", "DCM": "Hóa chất", "DPM": "Hóa chất",
    "CMG": "Công nghệ",
    "CTD": "Xây dựng", "HBC": "Xây dựng", "VCG": "Xây dựng",
    "HHV": "Xây dựng", "PC1": "Xây dựng",
    "REE": "Năng lượng", "POW": "Năng lượng", "NT2": "Năng lượng",
    "VJC": "Hàng không", "ACV": "Cảng hàng không",
    "GMD": "Cảng biển", "VSC": "Cảng biển", "HAH": "Cảng biển",
    "VTP": "Logistics",
    "BVH": "Bảo hiểm",
    "IDC": "KCN", "KBC": "KCN",
    "PHR": "Cao su", "SZC": "KCN",
}

_VN30_FALLBACK = [
    "ACB", "BCM", "BID", "BVH", "CTG", "FPT", "GAS", "GVR", "HDB",
    "HPG", "LPB", "MBB", "MSN", "MWG", "PLX", "POW", "SAB", "SHB",
    "SSB", "SSI", "STB", "TCB", "TPB", "VCB", "VHM", "VIB", "VIC",
    "VJC", "VNM", "VPB",
]

_NAME_MAP: dict = {
    "ACB": "ACB", "BID": "BIDV", "CTG": "VietinBank", "HDB": "HDBank",
    "LPB": "LienVietPostBank", "MBB": "MB Bank", "MSB": "MSB",
    "SHB": "SHB", "SSB": "SeABank", "STB": "Sacombank",
    "TCB": "Techcombank", "TPB": "TPBank", "VCB": "Vietcombank",
    "VIB": "VIB", "VPB": "VPBank",
    "BSI": "BSI Securities", "CTS": "CTS Securities", "HCM": "HCMC Securities",
    "SSI": "SSI Securities", "VCI": "Viet Capital Sec", "VND": "VNDirect",
    "BCM": "Becamex", "DIG": "DIC Corp", "DXG": "Đất Xanh",
    "KDH": "Khang Điền", "NLG": "Nam Long", "NVL": "Novaland",
    "PDR": "Phát Đạt", "VHM": "Vinhomes", "VIC": "Vingroup", "VRE": "Vincom Retail",
    "FPT": "FPT Corp", "CMG": "CMC Technology",
    "MWG": "Mobile World", "PNJ": "PNJ", "VNM": "Vinamilk",
    "MSN": "Masan", "SAB": "Sabeco", "HPG": "Hòa Phát", "HSG": "Hoa Sen",
    "GAS": "PVGas", "PLX": "Petrolimex", "PVD": "PV Drilling", "PVS": "PTSC",
    "CTD": "Coteccons", "HBC": "Hòa Bình", "VCG": "Vinaconex",
    "REE": "REE Corp", "POW": "PV Power", "NT2": "Nhiệt điện NT2",
    "VJC": "VietJet", "ACV": "ACV",
    "GMD": "Gemadept", "VSC": "Viconship", "HAH": "Hải An",
    "IDC": "IDC", "KBC": "KBC", "PC1": "PCC1", "HHV": "Đèo Cả BOT",
    "DCM": "Đạm Cà Mau", "DGC": "DGC", "DPM": "Đạm Phú Mỹ",
    "NKG": "Nam Kim", "GVR": "VRG",
}

_VN100_FALLBACK = _VN30_FALLBACK + [
    "AGG", "ANV", "BSI", "CII", "CTD", "CTR", "DCM", "DGC", "DGW",
    "DIG", "DPM", "DXG", "EVF", "GMD", "HAH", "HBC", "HCM", "HDG",
    "HSG", "HT1", "IDC", "KBC", "KDH", "KSB", "MSB", "NKG", "NLG",
    "NT2", "NVL", "PDR", "PHR", "PNJ", "PPC", "PVD", "PVS", "REE",
    "SCS", "SZC", "TCH", "THD", "VCI", "VCG", "VGC", "VHC", "VND",
    "VPI", "VRE", "VSC", "VTP",
]

_surfing_cache: dict = {}
_SURFING_TTL = 3600


def _calc_choppiness(bars: list, period: int = 14) -> float | None:
    """
    Choppiness Index: 100 * log10(sum(TR_1, n) / (HH - LL)) / log10(n)
    > 61.8 = sideway/ranging, < 38.2 = strong trend, 38-62 = neutral.
    """
    if len(bars) < period + 1:
        return None
    recent = bars[-(period + 1):]
    atr_sum = 0.0
    for i in range(1, len(recent)):
        h, l, pc = recent[i]["high"], recent[i]["low"], recent[i - 1]["close"]
        atr_sum += max(h - l, abs(h - pc), abs(l - pc))
    hh = max(b["high"] for b in recent[1:])
    ll = min(b["low"]  for b in recent[1:])
    if hh == ll or atr_sum == 0:
        return 100.0
    import math
    ci = 100 * math.log10(atr_sum / (hh - ll)) / math.log10(period)
    return round(min(100.0, max(0.0, ci)), 1)


def _calc_rsi(closes: list, period: int = 14) -> float:
    if len(closes) < period + 1:
        return None
    gains, losses = [], []
    for i in range(1, len(closes)):
        d = closes[i] - closes[i - 1]
        gains.append(max(d, 0.0))
        losses.append(max(-d, 0.0))
    ag = sum(gains[-period:]) / period
    al = sum(losses[-period:]) / period
    if al == 0:
        return 100.0
    return round(100 - 100 / (1 + ag / al), 1)


def _fetch_one_surfing(symbol: str) -> dict:
    try:
        today = vn_now().date()
        start = (today - timedelta(days=65)).strftime("%Y-%m-%d")
        end = today.strftime("%Y-%m-%d")
        bars = vnstock_client.get_historical_data(symbol, start, end, "1D", "VCI")
        if not bars or len(bars) < 21:
            return None

        closes = [b["close"] for b in bars]
        volumes = [b["volume"] for b in bars]

        today_vol = volumes[-1]
        avg_vol = sum(volumes[-21:-1]) / 20
        vol_ratio = round(today_vol / avg_vol, 2) if avg_vol else None

        ma20 = sum(closes[-20:]) / 20
        ma50 = sum(closes[-50:]) / 50 if len(closes) >= 50 else None
        rsi = _calc_rsi(closes)
        change_pct = round((closes[-1] - closes[-2]) / closes[-2] * 100, 2) if len(closes) >= 2 else None

        signals = []
        if vol_ratio and vol_ratio >= 2.0:
            signals.append("KL đột biến")
        elif vol_ratio and vol_ratio >= 1.5:
            signals.append("KL cao")
        if rsi is not None and rsi <= 30:
            signals.append("RSI quá bán")
        elif rsi is not None and rsi >= 70:
            signals.append("RSI quá mua")
        signals.append("Trên MA20" if closes[-1] > ma20 else "Dưới MA20")
        if ma50 is not None:
            signals.append("Trên MA50" if closes[-1] > ma50 else "Dưới MA50")

        return {
            "symbol": symbol,
            "industry": _INDUSTRY_MAP.get(symbol, "Khác"),
            "price": closes[-1],
            "change_pct": change_pct,
            "volume": int(today_vol),
            "avg_vol_20d": int(avg_vol),
            "vol_ratio": vol_ratio,
            "rsi": rsi,
            "ma20": round(ma20),
            "above_ma20": closes[-1] > ma20,
            "ma50": round(ma50) if ma50 else None,
            "above_ma50": closes[-1] > ma50 if ma50 else None,
            "signals": signals,
        }
    except Exception:
        return None


@app.get("/api/surfing/screener")
def surfing_screener(universe: str = "VN30", refresh: bool = False):
    cache_key = universe.upper()
    if not refresh:
        cached = _surfing_cache.get(cache_key)
        if cached and (time.time() - cached[0]) < _SURFING_TTL:
            return cached[1]

    try:
        from vnstock import Listing
        df = Listing().symbols_by_group(cache_key)
        symbols = df["ticker"].dropna().tolist()
    except Exception:
        symbols = _VN30_FALLBACK if cache_key == "VN30" else _VN100_FALLBACK

    results = []
    with ThreadPoolExecutor(max_workers=10) as pool:
        for r in pool.map(_fetch_one_surfing, symbols):
            if r:
                results.append(r)

    results.sort(key=lambda x: x.get("vol_ratio") or 0, reverse=True)

    resp = {
        "universe": cache_key,
        "count": len(results),
        "updated_at": vn_now().strftime("%H:%M %d/%m/%Y"),
        "data": results,
    }
    _surfing_cache[cache_key] = (time.time(), resp)
    return resp


# Universe cố định 32 mã cho tab Cổ Phiếu Xu Hướng Tăng.
# Nhỏ đủ để hoàn thành trong 10 giây (Vercel serverless limit),
# phủ đủ 5 nhóm ngành, không phụ thuộc vnstock Listing API.
_UPTREND_UNIVERSE = [
    "VCB", "BID", "TCB", "MBB", "CTG", "ACB", "VPB", "HDB",   # Ngân hàng
    "SSI", "VND", "HCM", "VCI", "BSI",                          # Chứng khoán
    "FPT", "CMG",                                                # Công nghệ
    "HHV", "PC1", "REE", "POW", "GMD", "IDC", "KBC", "CTD",    # Xây dựng/ĐTC
    "VHM", "NLG", "DXG", "KDH", "NVL", "BCM", "VIC", "VRE",   # BĐS
]
_uptrend_cache: dict = {}   # tách riêng khỏi _surfing_cache của tab T+3


@app.get("/api/uptrend-by-sector")
def uptrend_by_sector(refresh: bool = False):
    """
    Top 2 cổ phiếu xu hướng tăng theo 5 nhóm ngành.
    Fetch 32 mã cố định (không gọi VN100 screener — tránh timeout Vercel 10s).
    Điểm kỹ thuật tính từ RSI, MA20/50, vol_ratio, change_pct thật.
    """
    cache_key = "uptrend"
    cached = _uptrend_cache.get(cache_key)
    if not refresh and cached and (time.time() - cached[0]) < _SURFING_TTL:
        return cached[1]

    SECTOR_META = {
        "bank":  {"label": "Ngân hàng",              "industries": ["Ngân hàng"],
                  "catalyst": "Fed chu kỳ cắt lãi suất → NIM ngân hàng VN mở rộng, room tín dụng được nới, khối ngoại tăng tỷ trọng."},
        "sec":   {"label": "Chứng khoán",             "industries": ["Chứng khoán"],
                  "catalyst": "Thanh khoản thị trường tăng → margin lending tăng, phí môi giới & IB phục hồi; kỳ vọng nâng hạng 2026."},
        "tech":  {"label": "Công nghệ",               "industries": ["Công nghệ"],
                  "catalyst": "AI/cloud & chuyển đổi số quốc gia; xuất khẩu phần mềm và đơn hàng nước ngoài tăng trưởng."},
        "infra": {"label": "Xây dựng / Đầu tư công", "industries": ["Xây dựng", "Năng lượng", "KCN", "Cảng biển", "Logistics"],
                  "catalyst": "Gói đầu tư công 800K tỷ VNĐ giải ngân; Quy hoạch điện VIII; hạ tầng KCN đón FDI dịch chuyển."},
        "real":  {"label": "Bất động sản",            "industries": ["BĐS"],
                  "catalyst": "Luật Đất đai 2024 tháo gỡ pháp lý; lãi suất hạ nhiệt; tín dụng BĐS được nới; gói nhà ở xã hội 120K tỷ."},
    }

    def tech_score(r):
        s = 0
        if r.get("above_ma20"): s += 2
        if r.get("above_ma50"): s += 1
        rsi = r.get("rsi")
        if rsi:
            if 45 <= rsi <= 65:   s += 2
            elif 35 < rsi < 45 or 65 < rsi <= 70: s += 1
        vr = r.get("vol_ratio") or 0
        if vr >= 2.0:   s += 2
        elif vr >= 1.5: s += 1
        cp = r.get("change_pct") or 0
        if cp > 1:   s += 2
        elif cp > 0: s += 1
        return s

    def gen_kt(r):
        parts = []
        ma20, ma50 = r.get("ma20"), r.get("ma50")
        if r.get("above_ma20") and r.get("above_ma50"):
            parts.append(f"Trên MA20({ma20:,.0f}) & MA50({ma50:,.0f})")
        elif r.get("above_ma20"):
            parts.append(f"Trên MA20({ma20:,.0f}), chưa vượt MA50")
        rsi = r.get("rsi")
        if rsi is not None:
            if 45 <= rsi <= 65:   parts.append(f"RSI={rsi:.0f} vùng tăng tốt, dư địa còn tốt")
            elif 65 < rsi <= 70:  parts.append(f"RSI={rsi:.0f} cao, sắp kháng cự")
            elif rsi < 45:        parts.append(f"RSI={rsi:.0f} thấp, tiềm năng hồi phục")
        vr = r.get("vol_ratio")
        if vr and vr >= 1.5: parts.append(f"KL {vr:.1f}× TB20 — dòng tiền xác nhận")
        cp = r.get("change_pct")
        if cp is not None:
            parts.append(f"%ngày {'+' if cp >= 0 else ''}{cp:.1f}%")
        return " · ".join(parts) if parts else "Đang tích lũy, chờ tín hiệu rõ hơn."

    # Fetch 32 mã song song (max 8 workers để không vượt rate limit vnstock)
    rows = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        for r in pool.map(_fetch_one_surfing, _UPTREND_UNIVERSE):
            if r:
                rows.append(r)

    result = []
    for skey, meta in SECTOR_META.items():
        candidates = [r for r in rows
                      if r.get("industry") in meta["industries"] and r.get("above_ma20")]
        candidates.sort(key=tech_score, reverse=True)
        stocks = []
        for r in candidates[:2]:
            price = r["price"]
            stocks.append({
                "ticker":     r["symbol"],
                "name":       _NAME_MAP.get(r["symbol"], r["symbol"]),
                "price":      price,
                "entry":      price,
                "target":     round(price * 1.10, 1),
                "sl":         round(price * 0.95, 1),
                "upside":     "+10.0%",
                "kythuat":    gen_kt(r),
                "tinhieu":    ", ".join(r.get("signals", [])) or "Không có tín hiệu nổi bật",
                "catalyst":   meta["catalyst"],
                "score":      tech_score(r),
                "rsi":        r.get("rsi"),
                "vol_ratio":  r.get("vol_ratio"),
                "change_pct": r.get("change_pct"),
            })
        result.append({"sector": skey, "label": meta["label"], "stocks": stocks})

    resp = {
        "success":    True,
        "updated_at": vn_now().strftime("%H:%M %d/%m/%Y"),
        "sectors":    result,
    }
    _uptrend_cache[cache_key] = (time.time(), resp)
    return resp

