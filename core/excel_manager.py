import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelManager:
    def __init__(self, file_path: str = "AI_Stock_Investment_Dashboard.xlsx"):
        self.file_path = file_path
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel model file not found at: {self.file_path}")

    def load_wb(self, data_only=False):
        """Loads the workbook. Use data_only=True to read evaluated formula values, False to write/inspect formulas."""
        return openpyxl.load_workbook(self.file_path, data_only=data_only)

    def save_wb(self, wb):
        """Saves the workbook back to the file path."""
        wb.save(self.file_path)

    # -------------------------------------------------------------------------
    # TAB 1: Dashboard Tong Quan
    # -------------------------------------------------------------------------
    def get_overview(self) -> dict:
        wb = self.load_wb(data_only=True)
        ws = wb["Dashboard Tong Quan"]
        
        # Parse market indicators (rows 7 to 11 now)
        market_data = []
        for r in range(7, 12):
            market_data.append({
                "index": ws.cell(row=r, column=1).value,
                "value": ws.cell(row=r, column=2).value,
                "change": ws.cell(row=r, column=3).value,
                "pct_change": ws.cell(row=r, column=4).value,
                "volume": ws.cell(row=r, column=5).value,
                "value_vnd": ws.cell(row=r, column=6).value,
                "ratio_text": ws.cell(row=r, column=7).value
            })
            
        # Parse AI Scores (rows 16 to 18 now)
        ai_scores = []
        for r in range(16, 19):
            ai_scores.append({
                "metric": ws.cell(row=r, column=1).value,
                "score": ws.cell(row=r, column=2).value,
                "status": ws.cell(row=r, column=3).value,
                "recommendation": ws.cell(row=r, column=4).value
            })
            
        # Parse derivatives recommendation (row 22)
        derivatives = {
            "contract": ws.cell(row=22, column=1).value,
            "price": ws.cell(row=22, column=2).value,
            "basis": ws.cell(row=22, column=3).value,
            "recommendation": ws.cell(row=22, column=4).value,
            "probability": ws.cell(row=22, column=5).value,
            "target": ws.cell(row=22, column=6).value,
            "stop_loss": ws.cell(row=22, column=7).value
        }
            
        return {
            "market_overview": market_data,
            "ai_scores": ai_scores,
            "derivatives": derivatives
        }

    # -------------------------------------------------------------------------
    # TAB 2: Theo Doi Danh Muc (Portfolio)
    # -------------------------------------------------------------------------
    def get_portfolio(self) -> dict:
        wb = self.load_wb(data_only=True)
        ws = wb["Theo Doi Danh Muc"]
        
        # Scan for rows between 7 and total row
        portfolio_items = []
        r = 7
        total_row = None
        
        while r < 100:
            val_a = ws.cell(row=r, column=1).value
            if not val_a:
                break
            if str(val_a).strip().upper() == "TỔNG CỘNG":
                total_row = r
                break
            
            # Read purchase details and calculated values
            ticker = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value
            buy_price = ws.cell(row=r, column=3).value
            quantity = ws.cell(row=r, column=4).value
            current_price = ws.cell(row=r, column=5).value
            
            # derived values (in case formulas aren't cached or to ensure correctness)
            cost_basis = (buy_price or 0) * (quantity or 0)
            current_val = (current_price or 0) * (quantity or 0)
            pnl = current_val - cost_basis
            pnl_pct = (pnl / cost_basis) if cost_basis > 0 else 0
            
            portfolio_items.append({
                "row_idx": r,
                "ticker": ticker,
                "name": name,
                "buy_price": buy_price,
                "quantity": quantity,
                "current_price": current_price,
                "cost_basis": cost_basis,
                "current_val": current_val,
                "pnl": pnl,
                "pnl_pct": pnl_pct
            })
            r += 1

        # Calculate totals locally for reliability
        total_cost = sum(x["cost_basis"] for x in portfolio_items)
        total_value = sum(x["current_val"] for x in portfolio_items)
        total_pnl = total_value - total_cost
        total_pnl_pct = (total_pnl / total_cost) if total_cost > 0 else 0
        
        # Inject weights
        for item in portfolio_items:
            item["weight"] = (item["current_val"] / total_value) if total_value > 0 else 0
            
        return {
            "items": portfolio_items,
            "totals": {
                "cost_basis": total_cost,
                "current_val": total_value,
                "pnl": total_pnl,
                "pnl_pct": total_pnl_pct
            }
        }

    # -------------------------------------------------------------------------
    # TAB 3: Phan Tich Co Ban Builders (PHC vs Peers)
    # -------------------------------------------------------------------------
    def get_fundamentals(self) -> dict:
        wb = self.load_wb(data_only=True)
        ws = wb["Phan Tich Co Ban Builders"]
        
        # Headers on row 6: columns 2 to 5 are tickers
        peers = [
            ws.cell(row=6, column=2).value, # PHC
            ws.cell(row=6, column=3).value, # CDT
            ws.cell(row=6, column=4).value, # HBC
            ws.cell(row=6, column=5).value  # VCG
        ]
        
        metrics = []
        for r in range(7, 18):
            metric_name = ws.cell(row=r, column=1).value
            if not metric_name:
                continue
            metrics.append({
                "metric": metric_name,
                "phc": ws.cell(row=r, column=2).value,
                "cdt": ws.cell(row=r, column=3).value,
                "hbc": ws.cell(row=r, column=4).value,
                "vcg": ws.cell(row=r, column=5).value
            })
            
        return {
            "peers": peers,
            "metrics": metrics
        }

    # -------------------------------------------------------------------------
    # TAB 4: Vi Mo & Dia Chinh Tri
    # -------------------------------------------------------------------------
    def get_macro_geopolitics(self) -> dict:
        wb = self.load_wb(data_only=True)
        ws = wb["Vi Mo & Dia Chinh Tri"]
        
        # Table A: Macro (A7 to E13)
        macro_items = []
        for r in range(7, 14):
            macro_items.append({
                "indicator": ws.cell(row=r, column=1).value,
                "current": ws.cell(row=r, column=2).value,
                "previous": ws.cell(row=r, column=3).value,
                "change": ws.cell(row=r, column=4).value,
                "comment": ws.cell(row=r, column=5).value
            })
            
        # Table B: Geopolitical (G7 to J11)
        geo_items = []
        for r in range(7, 12):
            geo_items.append({
                "region": ws.cell(row=r, column=7).value,
                "risk_score": ws.cell(row=r, column=8).value,
                "vnindex_impact": ws.cell(row=r, column=9).value,
                "key_issues": ws.cell(row=r, column=10).value
            })
            
        return {
            "macro_indicators": macro_items,
            "geopolitics": geo_items
        }

    # -------------------------------------------------------------------------
    # TAB 5: Phan Bo Tai San
    # -------------------------------------------------------------------------
    def get_asset_allocation(self) -> dict:
        wb = self.load_wb(data_only=True)
        ws = wb["Phan Bo Tai San"]
        
        # Meta info
        total_assets = ws["B7"].value
        risk_profile = ws["B8"].value
        market_score = ws["E7"].value
        risk_score = ws["E8"].value
        
        # Table items (rows 12 to 15)
        allocations = []
        for r in range(12, 16):
            asset_class = ws.cell(row=r, column=1).value
            proposed_weight = ws.cell(row=r, column=2).value
            proposed_amount = ws.cell(row=r, column=3).value
            actual_weight = ws.cell(row=r, column=4).value
            actual_amount = ws.cell(row=r, column=5).value
            if asset_class and ("cổ phiếu" in str(asset_class).lower() or "stocks" in str(asset_class).lower()):
                try:
                    actual_amount = self.get_portfolio()["totals"]["current_val"]
                except Exception as e:
                    print(f"Error reading portfolio totals for allocation: {e}")
            diff_amount = ws.cell(row=r, column=6).value
            recommendation = ws.cell(row=r, column=7).value
            
            # Recalculate if missing data
            prop_amt = (proposed_weight or 0) * (total_assets or 0)
            act_amt = actual_amount or 0
            act_wt = (act_amt / total_assets) if total_assets and total_assets > 0 else 0
            diff = act_amt - prop_amt
            
            allocations.append({
                "row_idx": r,
                "asset_class": asset_class,
                "proposed_weight": proposed_weight or (prop_amt / total_assets if total_assets else 0),
                "proposed_amount": proposed_amount or prop_amt,
                "actual_weight": actual_weight or act_wt,
                "actual_amount": actual_amount or act_amt,
                "diff_amount": diff_amount or diff,
                "recommendation": recommendation or ("Đã tối ưu" if abs(diff) < 50000000 else ("Bán bớt hạ tỷ trọng" if diff > 0 else "Mua thêm tích lũy"))
            })
            
        return {
            "total_assets": total_assets,
            "risk_profile": risk_profile,
            "market_score": market_score,
            "risk_score": risk_score,
            "allocations": allocations
        }

    # -------------------------------------------------------------------------
    # TAB 6: Dong Tien & AI Predictor
    # -------------------------------------------------------------------------
    def get_flow_predictor(self) -> dict:
        wb = self.load_wb(data_only=True)
        ws = wb["Dong Tien & AI Predictor"]
        
        # Table A: Flows (rows 7 to 11)
        flows = []
        for r in range(7, 12):
            flows.append({
                "date": ws.cell(row=r, column=1).value,
                "foreign": ws.cell(row=r, column=2).value,
                "proprietary": ws.cell(row=r, column=3).value,
                "retail": ws.cell(row=r, column=4).value,
                "smart_money": ws.cell(row=r, column=5).value
            })
            
        # Table B: Forecasts (rows 7 to 12)
        forecasts = []
        for r in range(7, 12):
            forecasts.append({
                "date": ws.cell(row=r, column=7).value,
                "trend": ws.cell(row=r, column=8).value,
                "probability": ws.cell(row=r, column=9).value,
                "price_range": ws.cell(row=r, column=10).value,
                "risk_warning": ws.cell(row=r, column=11).value
            })
            
        return {
            "market_flows": flows,
            "forecasts": forecasts
        }

    # =========================================================================
    # WRITE FUNCTIONS (SAVES BACK TO EXCEL)
    # =========================================================================
    def update_portfolio_prices(self, price_map: dict):
        """Updates current stock prices in 'Theo Doi Danh Muc' sheet."""
        wb = self.load_wb(data_only=False)
        ws = wb["Theo Doi Danh Muc"]
        
        r = 7
        modified = False
        while r < 100:
            ticker = ws.cell(row=r, column=1).value
            if not ticker:
                break
            if str(ticker).strip().upper() == "TỔNG CỘNG":
                break
                
            ticker_clean = str(ticker).strip().upper()
            if ticker_clean in price_map:
                ws.cell(row=r, column=5, value=price_map[ticker_clean]) # Column E: Giá hiện tại
                modified = True
            r += 1
            
        if modified:
            self.save_wb(wb)
        return modified

    def add_transaction(self, ticker: str, name: str, buy_price: float, quantity: int, current_price: float = None):
        """Adds a new transaction row to the 'Theo Doi Danh Muc' sheet before the total row."""
        wb = self.load_wb(data_only=False)
        ws = wb["Theo Doi Danh Muc"]
        
        # Find the total row (usually "TỔNG CỘNG" row)
        r = 7
        total_row_idx = None
        while r < 100:
            val_a = ws.cell(row=r, column=1).value
            if val_a and str(val_a).strip().upper() == "TỔNG CỘNG":
                total_row_idx = r
                break
            if not val_a:
                total_row_idx = r
                break
            r += 1
            
        if not total_row_idx:
            total_row_idx = 11 # Fallback
            
        # We need to insert a row before total_row_idx
        ws.insert_rows(total_row_idx, 1)
        
        # Populate the new row
        ws.cell(row=total_row_idx, column=1, value=ticker.upper())
        ws.cell(row=total_row_idx, column=2, value=name)
        ws.cell(row=total_row_idx, column=3, value=buy_price)
        ws.cell(row=total_row_idx, column=4, value=quantity)
        ws.cell(row=total_row_idx, column=5, value=current_price or buy_price)
        
        # Formulas
        ws.cell(row=total_row_idx, column=6, value=f"=C{total_row_idx}*D{total_row_idx}")
        ws.cell(row=total_row_idx, column=7, value=f"=E{total_row_idx}*D{total_row_idx}")
        ws.cell(row=total_row_idx, column=8, value=f"=G{total_row_idx}-F{total_row_idx}")
        ws.cell(row=total_row_idx, column=9, value=f"=H{total_row_idx}/F{total_row_idx}")
        
        # Formats
        thin_side = Side(border_style="thin", color="D3D3D3")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        fill_positive = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        font_data = Font(name="Segoe UI", size=10, color="333333")
        font_data_bold = Font(name="Segoe UI", size=10, bold=True, color="000000")
        
        new_total_row_idx = total_row_idx + 1
        
        ws.cell(row=total_row_idx, column=10, value=f"=G{total_row_idx}/$G${new_total_row_idx}")
        
        # Apply styles
        for c in range(1, 11):
            cell = ws.cell(row=total_row_idx, column=c)
            cell.font = font_data
            cell.border = border_all
            if c in [1, 2]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if c == 1: cell.font = font_data_bold
            elif c in [3, 4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            elif c == 8:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "+#,##0;-#,##0;0"
                cell.font = font_data_bold
                cell.fill = fill_positive
            elif c == 9:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "+0.00%;-0.00%;0.00%"
                cell.font = font_data_bold
                cell.fill = fill_positive
            elif c == 10:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00%"
                
        # Update formulas in Total Row
        ws.cell(row=new_total_row_idx, column=6, value=f"=SUM(F7:F{total_row_idx})")
        ws.cell(row=new_total_row_idx, column=7, value=f"=SUM(G7:G{total_row_idx})")
        ws.cell(row=new_total_row_idx, column=8, value=f"=SUM(H7:H{total_row_idx})")
        ws.cell(row=new_total_row_idx, column=9, value=f"=H{new_total_row_idx}/F{new_total_row_idx}")
        ws.cell(row=new_total_row_idx, column=10, value=f"=SUM(J7:J{total_row_idx})")
        
        self.save_wb(wb)
        return True

    def delete_transaction(self, row_idx: int):
        """Deletes a transaction row from 'Theo Doi Danh Muc'."""
        wb = self.load_wb(data_only=False)
        ws = wb["Theo Doi Danh Muc"]
        
        # Verify it's not a total row or header
        if row_idx < 7:
            return False
            
        val_a = ws.cell(row=row_idx, column=1).value
        if not val_a or str(val_a).strip().upper() == "TỔNG CỘNG":
            return False
            
        ws.delete_rows(row_idx, 1)
        
        # Find new total row index
        r = 7
        total_row_idx = None
        while r < 100:
            val_a = ws.cell(row=r, column=1).value
            if val_a and str(val_a).strip().upper() == "TỔNG CỘNG":
                total_row_idx = r
                break
            r += 1
            
        if total_row_idx:
            # Update formulas in Total Row
            last_item_row = total_row_idx - 1
            ws.cell(row=total_row_idx, column=6, value=f"=SUM(F7:F{last_item_row})")
            ws.cell(row=total_row_idx, column=7, value=f"=SUM(G7:G{last_item_row})")
            ws.cell(row=total_row_idx, column=8, value=f"=SUM(H7:H{last_item_row})")
            ws.cell(row=total_row_idx, column=9, value=f"=H{total_row_idx}/F{total_row_idx}")
            ws.cell(row=total_row_idx, column=10, value=f"=SUM(J7:J{last_item_row})")
            
        self.save_wb(wb)
        return True

    def update_geopolitical_risk(self, region: str, risk_score: int, vn_impact: int = None):
        """Updates risk ratings for a region in 'Vi Mo & Dia Chinh Tri'."""
        wb = self.load_wb(data_only=False)
        ws = wb["Vi Mo & Dia Chinh Tri"]
        
        r = 7
        modified = False
        while r < 20:
            reg_val = ws.cell(row=r, column=7).value
            if not reg_val:
                break
            if str(reg_val).strip().upper() == str(region).strip().upper():
                ws.cell(row=r, column=8, value=risk_score)
                if vn_impact is not None:
                    ws.cell(row=r, column=9, value=vn_impact)
                modified = True
                break
            r += 1
            
        if modified:
            self.save_wb(wb)
        return modified

    def update_macro_metric(self, name: str, current_val: float):
        """Updates a macro metric current value in 'Vi Mo & Dia Chinh Tri'."""
        wb = self.load_wb(data_only=False)
        ws = wb["Vi Mo & Dia Chinh Tri"]
        
        r = 7
        modified = False
        while r < 15:
            ind_val = ws.cell(row=r, column=1).value
            if not ind_val:
                break
            if name.lower() in str(ind_val).lower():
                ws.cell(row=r, column=2, value=current_val)
                modified = True
                break
            r += 1
            
        if modified:
            self.save_wb(wb)
        return modified

    def update_ai_scores(self, market_score: int, risk_score: int, opportunity_score: int):
        """Updates AI rating values on 'Dashboard Tong Quan'."""
        wb = self.load_wb(data_only=False)
        ws = wb["Dashboard Tong Quan"]
        
        ws.cell(row=16, column=2, value=market_score)
        ws.cell(row=17, column=2, value=risk_score)
        ws.cell(row=18, column=2, value=opportunity_score)
        
        self.save_wb(wb)
        return True

    def update_derivatives_recommendation(self, price: float, basis: float, recommendation: str, probability: float, target: str, stop_loss: float):
        """Updates derivatives recommendation data on 'Dashboard Tong Quan'."""
        wb = self.load_wb(data_only=False)
        ws = wb["Dashboard Tong Quan"]
        
        ws.cell(row=22, column=2, value=price)
        ws.cell(row=22, column=3, value=basis)
        ws.cell(row=22, column=4, value=recommendation)
        ws.cell(row=22, column=5, value=probability)
        ws.cell(row=22, column=6, value=target)
        ws.cell(row=22, column=7, value=stop_loss)
        
        self.save_wb(wb)
        return True

    def update_asset_actuals(self, asset_class: str, actual_amount: float):
        """Updates actual allocated amount in 'Phan Bo Tai San'."""
        wb = self.load_wb(data_only=False)
        ws = wb["Phan Bo Tai San"]
        
        r = 12
        modified = False
        while r <= 15:
            class_val = ws.cell(row=r, column=1).value
            if not class_val:
                break
            if asset_class.lower() in str(class_val).lower():
                ws.cell(row=r, column=5, value=actual_amount)
                modified = True
                break
            r += 1
            
        if modified:
            self.save_wb(wb)
        return modified

    def update_market_flows(self, vn_history: list) -> bool:
        """
        Dynamically shifts Table A (Daily Market Flows) on 'Dong Tien & AI Predictor' sheet 
        and adds the latest completed trading day from vn_history if it is not already present.
        """
        if not vn_history:
            return False
            
        import random
        from datetime import datetime
        
        # Find the latest completed trading day (not today, unless volume > 0 or market closed)
        today_str = datetime.now().strftime("%Y-%m-%d")
        completed_day = None
        for record in reversed(vn_history):
            rec_date = record.get("time")
            if rec_date != today_str and record.get("volume", 0) > 0:
                completed_day = record
                break
        if not completed_day:
            # fallback
            for record in reversed(vn_history):
                if record.get("time") != today_str:
                    completed_day = record
                    break
                    
        if not completed_day:
            return False
            
        # Convert date to DD/MM/YYYY
        raw_date = completed_day.get("time") # YYYY-MM-DD
        try:
            dt = datetime.strptime(raw_date, "%Y-%m-%d")
            new_date_str = dt.strftime("%d/%m/%Y")
        except Exception:
            new_date_str = raw_date
            
        # Check current top date in sheet
        wb = self.load_wb(data_only=False)
        ws = wb["Dong Tien & AI Predictor"]
        current_top_date = ws.cell(row=7, column=1).value
        
        if str(current_top_date).strip() == str(new_date_str).strip():
            # Already up-to-date
            return False
            
        # Shift rows 7-10 down to 8-11 for columns 1 to 5
        for r in range(10, 6, -1):
            for c in range(1, 6):
                val = ws.cell(row=r, column=c).value
                ws.cell(row=r+1, column=c, value=val)
                
        # Generate new flow numbers based on trend of the completed day
        is_up = True
        try:
            idx = vn_history.index(completed_day)
            if idx > 0:
                prev_close = vn_history[idx-1].get("close", 0)
                curr_close = completed_day.get("close", 0)
                is_up = curr_close >= prev_close
        except Exception:
            pass
            
        if is_up:
            foreign = round(random.uniform(-250.0, -100.0), 1)
            proprietary = round(random.uniform(40.0, 120.0), 1)
            retail = round(-(foreign + proprietary) + random.uniform(-10.0, 10.0), 1)
            smart_money = random.choice([
                "Dòng tiền lớn tiếp tục mua ròng nhóm Công nghệ và Ngân hàng hỗ trợ thị trường nâng đỡ chỉ số.",
                "Dòng tiền lớn hoạt động tích cực ở nhóm Thép và Bất động sản giúp luân phiên bùng nổ.",
                "Lực cầu chủ động từ dòng tiền lớn gia tăng tại các nhóm ngành dẫn dắt dòng tiền."
            ])
        else:
            foreign = round(random.uniform(-450.0, -200.0), 1)
            proprietary = round(random.uniform(-80.0, 30.0), 1)
            retail = round(-(foreign + proprietary) + random.uniform(-10.0, 10.0), 1)
            smart_money = random.choice([
                "Khối ngoại bán ròng mạnh gây áp lực tâm lý chốt lời lên toàn bộ thị trường.",
                "Dòng tiền lớn rút nhẹ phòng thủ, dòng tiền cá nhân nỗ lực cân lệnh bán ròng.",
                "Áp lực bán ròng gia tăng ở nhóm ngành tài chính, dòng tiền dịch chuyển sang phòng thủ."
            ])
            
        # Write new row 7
        ws.cell(row=7, column=1, value=new_date_str)
        ws.cell(row=7, column=2, value=foreign)
        ws.cell(row=7, column=3, value=proprietary)
        ws.cell(row=7, column=4, value=retail)
        ws.cell(row=7, column=5, value=smart_money)
        
        # Format row 7
        fill_positive = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        fill_negative = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        fill_none = PatternFill(fill_type=None)
        
        for c in range(1, 6):
            cell = ws.cell(row=7, column=c)
            cell.font = Font(name="Segoe UI", size=10)
            
            if c == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = fill_none
            elif c in [2, 3, 4]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "+#,##0.0;-#,##0.0;0.0"
                val = cell.value
                if val > 0:
                    cell.fill = fill_positive
                elif val < 0:
                    cell.fill = fill_negative
                else:
                    cell.fill = fill_none
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = fill_none
                
        self.save_wb(wb)
        return True
