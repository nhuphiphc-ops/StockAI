import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load workbook
wb = openpyxl.load_workbook("AI_Stock_Investment_Dashboard.xlsx")
ws = wb["Dong Tien & AI Predictor"]

# Define new flows data for rows 7 to 11
# [date, foreign, proprietary, retail, smart_money]
new_flows = [
    ["12/06/2026", -185.4, 72.1, 113.3, "Dòng tiền lớn tiếp tục xoay vòng qua nhóm Thép và Bất động sản"],
    ["11/06/2026", -210.5, -15.2, 225.7, "Cá nhân hấp thụ tốt áp lực bán ròng từ khối ngoại"],
    ["10/06/2026", -152.4, -45.2, -107.2, "Dòng tiền lớn vào Thép và Ngân hàng"],
    ["09/06/2026", -82.1, 110.5, -28.4, "Dòng tiền lớn xoay vòng qua Khu công nghiệp"],
    ["08/06/2026", -124.5, 34.0, 90.5, "Cá nhân cân lệnh bán ròng của Khối ngoại"]
]

# Formatting elements
fill_positive = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid") # Soft Green
fill_negative = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid") # Soft Red
fill_none = PatternFill(fill_type=None)

for i, row_data in enumerate(new_flows):
    row_offset = 7 + i
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_offset, column=col_idx, value=val)
        
        # Style formatting (keep borders and Segoe UI font)
        cell.font = Font(name="Segoe UI", size=10)
        
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill_none
        elif col_idx in [2, 3, 4]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "+#,##0.0;-#,##0.0;0.0"
            if val > 0:
                cell.fill = fill_positive
            elif val < 0:
                cell.fill = fill_negative
            else:
                cell.fill = fill_none
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = fill_none

wb.save("AI_Stock_Investment_Dashboard.xlsx")
print("Successfully updated Flow Tracking Table (Table A) in Excel sheet.")
