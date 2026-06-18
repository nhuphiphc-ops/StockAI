import openpyxl

# Load workbook
wb = openpyxl.load_workbook("AI_Stock_Investment_Dashboard.xlsx")

# 1. Update Theo Doi Danh Muc sheet
ws_port = wb["Theo Doi Danh Muc"]

# Define new portfolio data:
# [ticker, name, buy_price, quantity, current_price]
new_portfolio = [
    ["FPT", "CTCP FPT", 80783, 4100, 73500],
    ["MBB", "CTCP Ngân hàng TMCP Quân đội", 26211, 3800, 24850],
    ["MBS", "CTCP Chứng khoán MB", 19600, 800, 19200],
    ["PHC", "CTCP Xây dựng Phục Hưng Holdings", 5675, 79000, 4640]
]

# Update rows 7 to 10
for i, item in enumerate(new_portfolio):
    row = 7 + i
    ws_port.cell(row=row, column=1, value=item[0]) # Mã CP
    ws_port.cell(row=row, column=2, value=item[1]) # Tên Doanh nghiệp
    ws_port.cell(row=row, column=3, value=item[2]) # Giá Mua (VND)
    ws_port.cell(row=row, column=4, value=item[3]) # Số lượng
    ws_port.cell(row=row, column=5, value=item[4]) # Giá hiện tại (VND)
    
    # Re-apply formulas to make sure they are correct
    ws_port.cell(row=row, column=6, value=f"=C{row}*D{row}") # Giá Vốn
    ws_port.cell(row=row, column=7, value=f"=E{row}*D{row}") # Giá trị Hiện tại
    ws_port.cell(row=row, column=8, value=f"=G{row}-F{row}") # Lãi / Lỗ VND
    ws_port.cell(row=row, column=9, value=f"=H{row}/F{row}") # Tỷ suất sinh lời
    ws_port.cell(row=row, column=10, value=f"=G{row}/$G$11") # Tỷ trọng

# Ensure the total row formulas are correct for these 4 rows (row 11)
ws_port.cell(row=11, column=1, value="TỔNG CỘNG")
ws_port.cell(row=11, column=6, value="=SUM(F7:F10)")
ws_port.cell(row=11, column=7, value="=SUM(G7:G10)")
ws_port.cell(row=11, column=8, value="=SUM(H7:H10)")
ws_port.cell(row=11, column=9, value="=H11/F11")
ws_port.cell(row=11, column=10, value="=SUM(J7:J10)")

# 2. Update Phan Bo Tai San sheet to link actual stock value to portfolio total
ws_alloc = wb["Phan Bo Tai San"]
ws_alloc.cell(row=12, column=5, value="='Theo Doi Danh Muc'!G11") # Link Actual Stocks amount

# Save workbook
wb.save("AI_Stock_Investment_Dashboard.xlsx")
print("Successfully updated Excel dashboard to match the portfolio in the image.")
